"""
Views for Assets app - Asset Management System.
Provides CRUD operations, search, filtering, and reporting for assets.
"""
import json

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin, UserPassesTestMixin
from django.contrib import messages
from django.urls import reverse_lazy, reverse
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.views.decorators.http import require_POST
from django.contrib.contenttypes.models import ContentType
from django.db.models import Q, Count, Sum
from django.http import JsonResponse, HttpResponse
from django.utils.translation import gettext_lazy as _
from django.core.paginator import Paginator
from django.core.exceptions import ObjectDoesNotExist
from django.db import transaction, IntegrityError
from django.db.models.deletion import ProtectedError
from django.conf import settings
import csv
import datetime
import pandas as pd
import io
from decimal import Decimal, InvalidOperation
from collections import OrderedDict

from .models import Asset, AssetCategory, AssetBrand, AssetModel, AssetAssignment, AssetMaintenance
from .forms import (
    AssetAssignmentForm,
    AssetBulkEditForm,
    AssetExportForm,
    AssetForm,
    AssetImportForm,
    AssetSearchForm,
    BrandForm,
    CategoryForm,
    ModelForm,
    hardware_brand_queryset,
    hardware_category_queryset,
    hardware_model_queryset,
)
from companies.models import Company, Division, Location, ImportRunChange
from audit.models import AuditLog
from utils.csv_import import read_csv_rows_with_fallback
from utils.import_rollback import (
    snapshot_instance,
    start_import_run,
    finalize_import_run,
    record_import_change,
    get_latest_rollback_run,
    rollback_run,
)


ASSETS_IMPORT_MODULE = 'assets'
ASSET_IMPORT_TYPE = 'asset'


def _assets_latest_rollback_url(request):
    latest_run = get_latest_rollback_run(request.user, ASSETS_IMPORT_MODULE, ASSET_IMPORT_TYPE)
    if latest_run is None:
        return None
    return reverse('assets:asset_import_rollback')


def _perform_assets_rollback(request):
    run = get_latest_rollback_run(request.user, ASSETS_IMPORT_MODULE, ASSET_IMPORT_TYPE)
    if run is None:
        messages.warning(request, _('No rollback-eligible import run was found.'))
        return redirect('assets:asset_import')

    outcome = rollback_run(run)
    if outcome['errors']:
        messages.warning(
            request,
            _('Rollback completed with warnings: %(count)s issue(s).') % {'count': len(outcome['errors'])}
        )
    messages.success(
        request,
        _('Rollback completed. Deleted %(deleted)s created records and restored %(restored)s updated records.') % {
            'deleted': outcome['deleted'],
            'restored': outcome['restored'],
        }
    )
    return redirect('assets:asset_import')


def _get_accessible_asset_log_queryset(user):
    asset_content_type = ContentType.objects.get_for_model(Asset)
    queryset = AuditLog.objects.select_related('user', 'content_type', 'company').prefetch_related('change_logs').filter(
        Q(content_type=asset_content_type) |
        Q(action__in=[AuditLog.ActionType.EXPORT, AuditLog.ActionType.IMPORT], description__icontains='asset')
    )

    if hasattr(user, 'is_superadmin') and user.is_superadmin():
        return queryset

    return queryset.filter(company__in=user.get_accessible_companies())


def _get_accessible_hardware_assets(user):
    return user.get_accessible_assets().exclude(category__item_type=AssetCategory.ItemType.SERVICE)


class AssetChangeLogListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """List asset-related audit log entries."""

    model = AuditLog
    template_name = 'audit/auditlog_list.html'
    context_object_name = 'audit_logs'
    paginate_by = 20

    def test_func(self):
        return self.request.user.can_view_audit()

    def get_queryset(self):
        queryset = _get_accessible_asset_log_queryset(self.request.user)
        search = (self.request.GET.get('search') or '').strip()
        if search:
            queryset = queryset.filter(
                Q(action__icontains=search) |
                Q(description__icontains=search) |
                Q(object_id__icontains=search) |
                Q(company__name__icontains=search) |
                Q(user__username__icontains=search) |
                Q(user__first_name__icontains=search) |
                Q(user__last_name__icontains=search)
            )
        return queryset.order_by('-timestamp')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Asset Change Logs')
        context['search'] = self.request.GET.get('search', '')
        context['back_url'] = reverse('assets:asset_list')
        return context


class AssetChangeLogDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    """Show one asset audit-log entry in detail."""

    model = AuditLog
    template_name = 'audit/auditlog_detail.html'
    context_object_name = 'audit_log'

    def test_func(self):
        return self.request.user.can_view_audit()

    def get_queryset(self):
        return _get_accessible_asset_log_queryset(self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        audit_log = self.object
        context['title'] = _('Asset Change Log Details')
        context['back_url'] = reverse('assets:asset_log_list')
        context['metadata_json'] = json.dumps(audit_log.metadata or {}, indent=2, ensure_ascii=False, default=str)

        related_asset = None
        if audit_log.content_type and audit_log.content_type.model == 'asset' and audit_log.object_id:
            related_asset = _get_accessible_hardware_assets(self.request.user).filter(pk=audit_log.object_id).first()

        context['related_asset'] = related_asset
        context['related_asset_url'] = reverse('assets:asset_detail', kwargs={'pk': related_asset.pk}) if related_asset else None
        return context


class AssetListView(LoginRequiredMixin, ListView):
    """List view for assets with filtering and search capabilities."""
    model = Asset
    template_name = 'assets/asset_list.html'
    context_object_name = 'assets'
    paginate_by = 25

    def _is_drilldown_mode(self):
        return bool((self.request.GET.get('drill_ids') or '').strip())

    def _build_base_queryset(self):
        queryset = _get_accessible_hardware_assets(self.request.user).select_related(
            'category', 'brand', 'model', 'company', 'assigned_to', 'location'
        )

        drill_ids = (self.request.GET.get('drill_ids') or '').strip()
        if drill_ids:
            ids = [value.strip() for value in drill_ids.split(',') if value.strip()]
            if ids:
                queryset = queryset.filter(pk__in=ids)

        # Apply filters
        search_query = self.request.GET.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(asset_number__icontains=search_query) |
                Q(description__icontains=search_query) |
                Q(serial_number__icontains=search_query) |
                Q(brand__name__icontains=search_query) |
                Q(model__name__icontains=search_query)
            )

        category = self.request.GET.get('category')
        if category:
            queryset = queryset.filter(category_id=category)

        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)

        location = self.request.GET.get('location')
        if location:
            queryset = queryset.filter(
                Q(location__name__icontains=location) |
                Q(location__code__icontains=location) |
                Q(current_location__icontains=location)
            )

        return queryset

    @staticmethod
    def _build_group_key(asset):
        legacy_location = (asset.current_location or '').strip().lower()
        return (
            asset.category_id,
            asset.brand_id,
            asset.model_id,
            asset.location_id,
            legacy_location,
            (asset.location_zone or '').strip().lower(),
            (asset.location_rack or '').strip().lower(),
            (asset.location_shelf or '').strip().lower(),
            asset.status,
            asset.assigned_to_id,
        )

    def _build_grouped_assets(self, queryset):
        grouped = OrderedDict()
        for asset in queryset:
            key = self._build_group_key(asset)
            if key not in grouped:
                grouped[key] = {
                    'representative': asset,
                    'quantity': 0,
                    'asset_ids': [],
                }
            grouped[key]['quantity'] += 1
            grouped[key]['asset_ids'].append(str(asset.pk))
        return list(grouped.values())

    @staticmethod
    def _build_location_slot_map(location_queryset):
        slot_map = {}
        for location in location_queryset:
            if location.location_type != Location.LocationType.WAREHOUSE:
                continue
            zones = location.expanded_zones()
            racks = location.expanded_racks()
            shelves = location.expanded_shelves()
            if not zones or not racks or not shelves:
                continue
            slot_map[str(location.pk)] = {
                'zones': zones,
                'racks': racks,
                'shelves': shelves,
            }
        return slot_map
    
    def get_queryset(self):
        base_queryset = self._build_base_queryset()
        self._base_queryset = base_queryset
        self._is_drilldown = self._is_drilldown_mode()

        if self._is_drilldown:
            queryset = base_queryset
        else:
            queryset = base_queryset.exclude(Q(serial_number__isnull=True) | Q(serial_number=''))

        # Ordering
        order_by = self.request.GET.get('order_by', '-created_at')
        queryset = queryset.order_by(order_by)

        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categories'] = hardware_category_queryset().order_by('name')
        context['locations'] = Location.objects.filter(status='active')
        context['search_form'] = AssetSearchForm(self.request.GET)
        context['status_choices'] = Asset.AssetStatus.choices
        context['is_drilldown'] = getattr(self, '_is_drilldown', self._is_drilldown_mode())
        context['current_query'] = self.request.GET.urlencode()
        context['bulk_edit_form'] = AssetBulkEditForm(user=self.request.user)
        context['location_slot_map'] = self._build_location_slot_map(
            context['bulk_edit_form'].fields['location'].queryset
        )

        base_queryset = getattr(self, '_base_queryset', self._build_base_queryset())
        if context['is_drilldown']:
            context['grouped_assets'] = []
        else:
            grouped_queryset = base_queryset.filter(Q(serial_number__isnull=True) | Q(serial_number='')).order_by('-created_at')
            context['grouped_assets'] = self._build_grouped_assets(grouped_queryset)

        # Stats for dashboard cards
        context['total_assets'] = base_queryset.count()
        context['available_assets'] = base_queryset.filter(status='available').count()
        context['assigned_assets'] = base_queryset.filter(status='assigned').count()
        context['maintenance_assets'] = base_queryset.filter(status='maintenance').count()

        return context


@login_required
def asset_bulk_edit_view(request):
    """Apply selected field updates to a quantity of selected assets."""
    if request.method != 'POST':
        return redirect('assets:asset_list')

    form = AssetBulkEditForm(request.POST, user=request.user)
    return_query = (request.POST.get('return_query') or '').strip()

    def _redirect_to_list():
        base_url = reverse('assets:asset_list')
        if return_query:
            return redirect(f"{base_url}?{return_query}")
        return redirect(base_url)

    if not form.is_valid():
        first_error = '; '.join(form.non_field_errors())
        if not first_error:
            for errors in form.errors.values():
                if errors:
                    first_error = errors[0]
                    break
        messages.error(request, first_error or _('Bulk update failed. Please check the input values.'))
        return _redirect_to_list()

    selected_ids = form.cleaned_data['asset_ids']
    accessible_assets = _get_accessible_hardware_assets(request.user).filter(pk__in=selected_ids)
    asset_map = {str(asset.pk): asset for asset in accessible_assets}
    ordered_assets = [asset_map[asset_id] for asset_id in selected_ids if asset_id in asset_map]

    if not ordered_assets:
        messages.error(request, _('No selected assets are available for update.'))
        return _redirect_to_list()

    apply_quantity = form.cleaned_data.get('apply_quantity') or len(ordered_assets)
    apply_quantity = min(apply_quantity, len(ordered_assets))
    target_assets = ordered_assets[:apply_quantity]
    update_data = form.get_update_data()

    updated_count = 0
    with transaction.atomic():
        for asset in target_assets:
            changes = []
            for field, new_value in update_data.items():
                old_value = getattr(asset, field)
                if old_value != new_value:
                    setattr(asset, field, new_value)
                    changes.append(f'{field}: {old_value} -> {new_value}')

            if not changes:
                continue

            asset.save()
            updated_count += 1

            AuditLog.objects.create(
                user=request.user,
                company=asset.company,
                action=AuditLog.ActionType.UPDATE,
                content_object=asset,
                description=f'Bulk updated asset: {asset.asset_number}. Changes: {", ".join(changes)}',
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )

    if updated_count:
        messages.success(request, _('Successfully updated %(count)s assets.') % {'count': updated_count})
    else:
        messages.info(request, _('No assets changed.'))

    return _redirect_to_list()


class AssetDetailView(LoginRequiredMixin, DetailView):
    """Detailed view of a single asset."""
    model = Asset
    template_name = 'assets/asset_detail.html'
    context_object_name = 'asset'
    
    def get_object(self):
        accessible_assets = _get_accessible_hardware_assets(self.request.user)
        obj = get_object_or_404(accessible_assets, pk=self.kwargs['pk'])
        return obj
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        asset = self.get_object()
        
        # Get assignment history
        context['assignments'] = AssetAssignment.objects.filter(
            asset=asset
        ).select_related('assigned_to', 'assigned_by').order_by('-assigned_date')
        
        # Get maintenance history
        context['maintenance_records'] = AssetMaintenance.objects.filter(
            asset=asset
        ).order_by('-scheduled_date')
        
        # Get audit logs for this asset
        context['audit_logs'] = AuditLog.objects.filter(
            object_id=str(asset.pk),
            content_type__model='asset'
        ).select_related('user').order_by('-timestamp')[:10]
        
        return context


class AssetCreateView(LoginRequiredMixin, CreateView):
    """Create new asset."""
    model = Asset
    form_class = AssetForm
    template_name = 'assets/asset_form.html'
    success_url = reverse_lazy('assets:asset_list')

    def _resolve_target_company(self, cleaned_data):
        """Resolve company for new assets from user context, then location fallback."""
        user_company = None
        try:
            user_company = getattr(self.request.user, 'company', None)
        except ObjectDoesNotExist:
            user_company = None

        if user_company:
            return user_company

        location = cleaned_data.get('location')
        if location and getattr(location, 'company', None):
            return location.company

        return None

    def _build_model_catalog(self):
        models = hardware_model_queryset().select_related('brand', 'category').order_by('brand__name', 'name')
        return [
            {
                'id': str(model.pk),
                'name': model.name,
                'brand_id': str(model.brand_id) if model.brand_id else '',
                'brand_name': model.brand.name if model.brand_id else '',
                'category_id': str(model.category_id) if model.category_id else '',
                'label': f"{model.brand.name} - {model.name}" if model.brand_id else model.name,
            }
            for model in models
        ]

    def _validate_batch_rows(self, form, amount, brand):
        form._duplicate_serial_rows = []
        serial_numbers = self.request.POST.getlist('batch_serial_number[]')
        statuses = self.request.POST.getlist('batch_status[]')
        valid_statuses = {choice[0] for choice in Asset.AssetStatus.choices}

        raw_normalized_rows = []
        for idx in range(amount):
            raw_serial = serial_numbers[idx] if idx < len(serial_numbers) else ''
            normalized = (raw_serial or '').strip().lower()
            if normalized:
                raw_normalized_rows.append((idx + 1, normalized))

        seen_raw = {}
        for row_idx, normalized_serial in raw_normalized_rows:
            first_row = seen_raw.get(normalized_serial)
            if first_row is not None:
                form._duplicate_serial_rows = [first_row, row_idx]
                form.add_error(
                    None,
                    _('Duplicate serial number detected in rows %(row1)s and %(row2)s. Serial numbers must be unique within the same brand regardless of status.')
                    % {
                        'row1': first_row,
                        'row2': row_idx,
                    }
                )
                return None, None
            seen_raw[normalized_serial] = row_idx

        if len(statuses) < amount:
            form.add_error(None, _('Please provide status for each item in the batch.'))
            return None, None

        parsed_rows = []
        for idx in range(amount):
            row_status = (statuses[idx] if idx < len(statuses) else '').strip()
            row_serial = (serial_numbers[idx] if idx < len(serial_numbers) else '').strip()
            if row_status not in valid_statuses:
                form.add_error(None, _('Invalid status at row %(row)s.') % {'row': idx + 1})
                return None, None
            raw_serial = serial_numbers[idx] if idx < len(serial_numbers) else ''
            if (raw_serial or '').strip() and not row_serial:
                form._duplicate_serial_rows = [idx + 1]
                form.add_error(
                    None,
                    _('Serial number at row %(row)s is invalid. Please re-enter it manually.')
                    % {'row': idx + 1}
                )
                return None, None
            parsed_rows.append((row_serial, row_status))

        if not brand:
            return parsed_rows, serial_numbers

        normalized_rows = []
        for idx, (row_serial, _row_status) in enumerate(parsed_rows, start=1):
            normalized = row_serial.strip().lower()
            if normalized:
                normalized_rows.append((idx, row_serial.strip(), normalized))

        seen_in_batch = {}
        for row_idx, raw_serial, normalized_serial in normalized_rows:
            first_row = seen_in_batch.get(normalized_serial)
            if first_row is not None:
                form._duplicate_serial_rows = [first_row, row_idx]
                form.add_error(
                    None,
                    _('Serial number "%(serial)s" is duplicated for brand %(brand)s in rows %(row1)s and %(row2)s.')
                    % {
                        'serial': raw_serial,
                        'brand': brand.name,
                        'row1': first_row,
                        'row2': row_idx,
                    }
                )
                return None, None
            seen_in_batch[normalized_serial] = row_idx

        existing_serials = {
            (value or '').strip().lower()
            for value in Asset.objects.filter(brand=brand)
            .exclude(serial_number='')
            .exclude(serial_number__isnull=True)
            .values_list('serial_number', flat=True)
            if (value or '').strip()
        }
        for row_idx, raw_serial, normalized_serial in normalized_rows:
            if normalized_serial in existing_serials:
                form._duplicate_serial_rows = [row_idx]
                form.add_error(
                    None,
                    _('Serial number "%(serial)s" already exists under brand %(brand)s (row %(row)s).')
                    % {
                        'serial': raw_serial,
                        'brand': brand.name,
                        'row': row_idx,
                    }
                )
                return None, None

        return parsed_rows, serial_numbers

    def _create_batch_assets(self, form, rows, target_company):
        cleaned = form.cleaned_data
        created_assets = []

        for row_serial, row_status in rows:
            asset = Asset(
                asset_number='',  # always auto-generate for batch to avoid unique collisions
                category=cleaned.get('category'),
                brand=cleaned.get('brand'),
                model=cleaned.get('model'),
                serial_number=row_serial,
                description=cleaned.get('description', ''),
                location=cleaned.get('location'),
                location_zone=cleaned.get('location_zone'),
                location_rack=cleaned.get('location_rack'),
                location_shelf=cleaned.get('location_shelf'),
                status=row_status,
                purchase_date=cleaned.get('purchase_date'),
                purchase_price=cleaned.get('purchase_price'),
                warranty_provider=cleaned.get('warranty_provider', ''),
                warranty_end_date=cleaned.get('warranty_end_date'),
                notes=cleaned.get('notes', ''),
                photo=cleaned.get('photo'),
                company=target_company,
                created_by=self.request.user,
            )
            asset.save()
            created_assets.append(asset)

            AuditLog.objects.create(
                user=self.request.user,
                company=target_company,
                action=AuditLog.ActionType.CREATE,
                content_object=asset,
                description=f'Created asset: {asset.asset_number}',
                ip_address=self.request.META.get('REMOTE_ADDR'),
                user_agent=self.request.META.get('HTTP_USER_AGENT', '')
            )

        return created_assets
    
    def form_valid(self, form):
        target_company = self._resolve_target_company(form.cleaned_data)
        if not target_company:
            form.add_error(None, _('Unable to determine company for this asset. Please choose a location linked to a company.'))
            return self.form_invalid(form)

        amount = form.cleaned_data.get('amount') or 1
        amount = max(1, int(amount))

        rows, _raw_serials = self._validate_batch_rows(form, amount, form.cleaned_data.get('brand'))
        if rows is None:
            return self.form_invalid(form)

        # Keep existing single-create path for backward compatibility.
        if amount == 1:
            first_serial, first_status = rows[0]
            form.instance.company = target_company
            form.instance.created_by = self.request.user
            form.instance.serial_number = first_serial
            form.instance.status = first_status

            try:
                with transaction.atomic():
                    response = super().form_valid(form)

                    AuditLog.objects.create(
                        user=self.request.user,
                        company=target_company,
                        action=AuditLog.ActionType.CREATE,
                        content_object=self.object,
                        description=f'Created asset: {self.object.asset_number}',
                        ip_address=self.request.META.get('REMOTE_ADDR'),
                        user_agent=self.request.META.get('HTTP_USER_AGENT', '')
                    )

                    messages.success(
                        self.request,
                        _('Asset "{}" has been created successfully.').format(self.object.asset_number)
                    )

                    return response
            except IntegrityError:
                form.add_error(None, _('Unable to create asset because of duplicate unique fields. Check serial number/barcode for this brand and try again.'))
                return self.form_invalid(form)

        if form.cleaned_data.get('asset_number'):
            form.add_error('asset_number', _('Leave asset number blank when creating multiple assets.'))
            return self.form_invalid(form)

        try:
            with transaction.atomic():
                created_assets = self._create_batch_assets(form, rows, target_company)
        except IntegrityError:
            form.add_error(None, _('Batch creation failed because one or more serial number/barcode values are duplicated. Serial numbers must be unique within each brand.'))
            return self.form_invalid(form)

        messages.success(
            self.request,
            _('Successfully created %(count)s assets in batch.') % {'count': len(created_assets)}
        )
        return redirect(self.success_url)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['model_catalog'] = self._build_model_catalog()
        context['asset_status_choices'] = Asset.AssetStatus.choices
        form = context.get('form')
        context['location_slot_map'] = form.get_location_slot_map() if form else {}
        context['duplicate_serial_rows'] = getattr(form, '_duplicate_serial_rows', []) if form else []
        return context


class AssetUpdateView(LoginRequiredMixin, UpdateView):
    """Update existing asset."""
    model = Asset
    form_class = AssetForm
    template_name = 'assets/asset_form.html'
    
    def get_object(self):
        accessible_assets = _get_accessible_hardware_assets(self.request.user)
        return get_object_or_404(accessible_assets, pk=self.kwargs['pk'])
    
    def form_valid(self, form):
        with transaction.atomic():
            # Store original values for audit log
            original_asset = Asset.objects.get(pk=self.object.pk)
            changes = []
            
            for field in form.changed_data:
                old_value = getattr(original_asset, field)
                new_value = form.cleaned_data[field]
                changes.append(f'{field}: {old_value} → {new_value}')
            
            response = super().form_valid(form)
            
            # Log the update
            if changes:
                AuditLog.objects.create(
                    user=self.request.user,
                    company=self.request.user.company,
                    action=AuditLog.ActionType.UPDATE,
                    content_object=self.object,
                    description=f'Updated asset: {self.object.asset_number}. Changes: {", ".join(changes)}',
                    ip_address=self.request.META.get('REMOTE_ADDR'),
                    user_agent=self.request.META.get('HTTP_USER_AGENT', '')
                )
            
            messages.success(
                self.request,
                _('Asset "{}" has been updated successfully.').format(self.object.asset_number)
            )
            
            return response
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def get_success_url(self):
        return reverse('assets:asset_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        models = hardware_model_queryset().select_related('brand', 'category').order_by('brand__name', 'name')
        context['model_catalog'] = [
            {
                'id': str(model.pk),
                'name': model.name,
                'brand_id': str(model.brand_id) if model.brand_id else '',
                'brand_name': model.brand.name if model.brand_id else '',
                'category_id': str(model.category_id) if model.category_id else '',
                'label': f"{model.brand.name} - {model.name}" if model.brand_id else model.name,
            }
            for model in models
        ]
        form = context.get('form')
        context['location_slot_map'] = form.get_location_slot_map() if form else {}
        return context


class AssetDeleteView(LoginRequiredMixin, DeleteView):
    """Delete asset (soft delete)."""
    model = Asset
    template_name = 'assets/asset_delete.html'
    success_url = reverse_lazy('assets:asset_list')
    
    def get_object(self):
        accessible_assets = _get_accessible_hardware_assets(self.request.user)
        return get_object_or_404(accessible_assets, pk=self.kwargs['pk'])
    
    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        
        with transaction.atomic():
            # Soft delete - mark as retired instead of actual deletion
            self.object.status = 'retired'
            self.object.save()
            
            # Log the deletion
            AuditLog.objects.create(
                user=request.user,
                company=request.user.company,
                action=AuditLog.ActionType.DELETE,
                content_object=self.object,
                description=f'Deleted (retired) asset: {self.object.asset_number}',
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            
            messages.success(
                request,
                _('Asset "{}" has been retired successfully.').format(self.object.asset_number)
            )
        
        return redirect(self.success_url)


@login_required
def asset_assign_view(request, pk):
    """Assign asset to a user."""
    accessible_assets = _get_accessible_hardware_assets(request.user)
    asset = get_object_or_404(accessible_assets, pk=pk)
    
    if request.method == 'POST':
        form = AssetAssignmentForm(request.POST, user=request.user)
        if form.is_valid():
            with transaction.atomic():
                # Return current assignment if exists
                current_assignment = AssetAssignment.objects.filter(
                    asset=asset, returned_date__isnull=True
                ).first()
                
                if current_assignment:
                    current_assignment.returned_date = datetime.datetime.now()
                    current_assignment.returned_by = request.user
                    current_assignment.save()
                
                # Create new assignment
                assignment = form.save(commit=False)
                assignment.asset = asset
                assignment.assigned_by = request.user
                assignment.save()
                
                # Update asset status
                asset.status = 'assigned'
                asset.assigned_to = assignment.assigned_to
                asset.save()
                
                # Log the assignment
                AuditLog.objects.create(
                    user=request.user,
                    company=request.user.company,
                    action=AuditLog.ActionType.ASSIGN,
                    content_object=asset,
                    description=f'Assigned asset {asset.asset_number} to {assignment.assigned_to.get_display_name()}',
                    ip_address=request.META.get('REMOTE_ADDR'),
                    user_agent=request.META.get('HTTP_USER_AGENT', '')
                )
                
                messages.success(
                    request,
                    _('Asset "{}" has been assigned to {}.').format(
                        asset.asset_number, assignment.assigned_to.get_display_name()
                    )
                )
                
                return redirect('assets:asset_detail', pk=asset.pk)
    else:
        form = AssetAssignmentForm(user=request.user)
    
    return render(request, 'assets/asset_assign.html', {
        'asset': asset,
        'form': form
    })


@login_required
def asset_return_view(request, pk):
    """Return asset from current assignment."""
    accessible_assets = _get_accessible_hardware_assets(request.user)
    asset = get_object_or_404(accessible_assets, pk=pk)
    
    current_assignment = AssetAssignment.objects.filter(
        asset=asset, returned_at__isnull=True
    ).first()
    
    if not current_assignment:
        messages.error(request, _('This asset is not currently assigned.'))
        return redirect('assets:asset_detail', pk=asset.pk)
    
    if request.method == 'POST':
        with transaction.atomic():
            current_assignment.returned_date = datetime.datetime.now()
            current_assignment.returned_by = request.user
            current_assignment.return_notes = request.POST.get('return_notes', '')
            current_assignment.save()
            
            # Update asset status
            asset.status = 'available'
            asset.assigned_to = None
            asset.save()
            
            # Log the return
            AuditLog.objects.create(
                user=request.user,
                company=request.user.company,
                action=AuditLog.ActionType.RETURN,
                content_object=asset,
                description=f'Returned asset {asset.asset_number} from {current_assignment.assigned_to.get_display_name()}',
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            
            messages.success(
                request,
                _('Asset "{}" has been returned successfully.').format(asset.asset_number)
            )
        
        return redirect('assets:asset_detail', pk=asset.pk)
    
    return render(request, 'assets/asset_return.html', {
        'asset': asset,
        'assignment': current_assignment
    })


@login_required
def asset_export_view(request):
    """View for asset export with filters and format selection."""
    accessible_assets = _get_accessible_hardware_assets(request.user).select_related(
        'category', 'brand', 'model', 'assigned_to', 'location'
    )

    if request.method == 'POST':
        form = AssetExportForm(user=request.user, data=request.POST)
        if form.is_valid():
            # Get base queryset
            base_queryset = accessible_assets
            
            # Apply filters
            filtered_queryset = form.get_filtered_queryset(base_queryset)
            
            # Get export format and fields
            export_format = form.cleaned_data['export_format']
            include_fields = form.cleaned_data['include_fields']
            
            # Generate appropriate export
            if export_format == 'csv':
                return generate_csv_export(request, filtered_queryset, include_fields)
            elif export_format == 'excel':
                return generate_excel_export(request, filtered_queryset, include_fields)
            elif export_format == 'pdf':
                return generate_pdf_export(request, filtered_queryset, include_fields)
    else:
        form = AssetExportForm(user=request.user)
    
    # Get stats for display
    total_assets = accessible_assets.count()
    
    context = {
        'form': form,
        'total_assets': total_assets,
    }
    return render(request, 'assets/asset_export.html', context)


def generate_csv_export(request, queryset, include_fields):
    """Generate CSV export with selected fields."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="assets_export_{datetime.date.today()}.csv"'
    
    writer = csv.writer(response)
    
    # Define field mapping
    field_mapping = {
        'asset_number': ('Asset Number', lambda asset: asset.asset_number),
        'category': ('Category', lambda asset: asset.category.name if asset.category else ''),
        'brand': ('Brand', lambda asset: asset.brand.name if asset.brand else ''),
        'model': ('Model', lambda asset: asset.model.name if asset.model else ''),
        'serial_number': ('Serial Number', lambda asset: asset.serial_number or ''),
        'description': ('Description', lambda asset: asset.description or ''),
        'status': ('Status', lambda asset: asset.get_status_display()),
        'location': ('Location', lambda asset: asset.get_location_display_abbr()),
        'assigned_to': ('Assigned To', lambda asset: asset.assigned_to.get_display_name() if asset.assigned_to else ''),
        'purchase_date': ('Purchase Date', lambda asset: asset.purchase_date.strftime('%Y-%m-%d') if asset.purchase_date else ''),
        'purchase_price': ('Purchase Price', lambda asset: str(asset.purchase_price) if asset.purchase_price else ''),
        'warranty_end_date': ('Warranty End Date', lambda asset: asset.warranty_end_date.strftime('%Y-%m-%d') if asset.warranty_end_date else ''),
        'created_at': ('Created At', lambda asset: asset.created_at.strftime('%Y-%m-%d %H:%M:%S')),
        'updated_at': ('Updated At', lambda asset: asset.updated_at.strftime('%Y-%m-%d %H:%M:%S')),
    }
    
    # Write header
    headers = [field_mapping[field][0] for field in include_fields if field in field_mapping]
    writer.writerow(headers)
    
    # Write data
    for asset in queryset:
        row = [field_mapping[field][1](asset) for field in include_fields if field in field_mapping]
        writer.writerow(row)
    
    # Log the export
    AuditLog.objects.create(
        user=request.user,
        company=request.user.company,
        action=AuditLog.ActionType.EXPORT,
        description=f'Exported {queryset.count()} assets to CSV with filters',
        ip_address=request.META.get('REMOTE_ADDR'),
        user_agent=request.META.get('HTTP_USER_AGENT', '')
    )
    
    return response


def generate_excel_export(request, queryset, include_fields):
    """Generate Excel export with selected fields."""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        messages.error(request, _('Excel export requires openpyxl package. Please contact administrator.'))
        return redirect('assets:asset_export')
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assets Export"
    
    # Define field mapping (same as CSV)
    field_mapping = {
        'asset_number': ('Asset Number', lambda asset: asset.asset_number),
        'category': ('Category', lambda asset: asset.category.name if asset.category else ''),
        'brand': ('Brand', lambda asset: asset.brand.name if asset.brand else ''),
        'model': ('Model', lambda asset: asset.model.name if asset.model else ''),
        'serial_number': ('Serial Number', lambda asset: asset.serial_number or ''),
        'description': ('Description', lambda asset: asset.description or ''),
        'status': ('Status', lambda asset: asset.get_status_display()),
        'location': ('Location', lambda asset: asset.get_location_display_abbr()),
        'assigned_to': ('Assigned To', lambda asset: asset.assigned_to.get_display_name() if asset.assigned_to else ''),
        'purchase_date': ('Purchase Date', lambda asset: asset.purchase_date.strftime('%Y-%m-%d') if asset.purchase_date else ''),
        'purchase_price': ('Purchase Price', lambda asset: str(asset.purchase_price) if asset.purchase_price else ''),
        'warranty_end_date': ('Warranty End Date', lambda asset: asset.warranty_end_date.strftime('%Y-%m-%d') if asset.warranty_end_date else ''),
        'created_at': ('Created At', lambda asset: asset.created_at.strftime('%Y-%m-%d %H:%M:%S')),
        'updated_at': ('Updated At', lambda asset: asset.updated_at.strftime('%Y-%m-%d %H:%M:%S')),
    }
    
    # Write headers with formatting
    headers = [field_mapping[field][0] for field in include_fields if field in field_mapping]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Write data
    for row, asset in enumerate(queryset, 2):
        for col, field in enumerate(include_fields, 1):
            if field in field_mapping:
                value = field_mapping[field][1](asset)
                ws.cell(row=row, column=col, value=value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="assets_export_{datetime.date.today()}.xlsx"'
    
    wb.save(response)
    
    # Log the export
    AuditLog.objects.create(
        user=request.user,
        company=request.user.company,
        action=AuditLog.ActionType.EXPORT,
        description=f'Exported {queryset.count()} assets to Excel with filters',
        ip_address=request.META.get('REMOTE_ADDR'),
        user_agent=request.META.get('HTTP_USER_AGENT', '')
    )
    
    return response


def generate_pdf_export(request, queryset, include_fields):
    """Generate PDF export with selected fields."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        import io
    except ImportError:
        messages.error(request, _('PDF export requires reportlab package. Please contact administrator.'))
        return redirect('assets:asset_export')
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Center
    )
    
    # Content
    content = []
    
    # Title
    title = Paragraph(f"Asset Export Report - {datetime.date.today()}", title_style)
    content.append(title)
    content.append(Spacer(1, 12))
    
    # Summary
    summary = Paragraph(f"Total Assets: {queryset.count()}", styles['Normal'])
    content.append(summary)
    content.append(Spacer(1, 12))
    
    # Field mapping
    field_mapping = {
        'asset_number': ('Asset #', lambda asset: asset.asset_number),
        'category': ('Category', lambda asset: asset.category.name if asset.category else ''),
        'brand': ('Brand', lambda asset: asset.brand.name if asset.brand else ''),
        'model': ('Model', lambda asset: asset.model.name if asset.model else ''),
        'serial_number': ('Serial #', lambda asset: asset.serial_number or ''),
        'status': ('Status', lambda asset: asset.get_status_display()),
        'location': ('Location', lambda asset: asset.get_location_display_abbr()),
        'assigned_to': ('Assigned To', lambda asset: asset.assigned_to.get_display_name() if asset.assigned_to else ''),
        'purchase_date': ('Purchase Date', lambda asset: asset.purchase_date.strftime('%Y-%m-%d') if asset.purchase_date else ''),
    }
    
    # Create table data
    headers = [field_mapping[field][0] for field in include_fields if field in field_mapping]
    table_data = [headers]
    
    for asset in queryset[:100]:  # Limit for PDF readability
        row = [field_mapping[field][1](asset) for field in include_fields if field in field_mapping]
        table_data.append(row)
    
    # Create table
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    content.append(table)
    
    # Build PDF
    doc.build(content)
    
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="assets_export_{datetime.date.today()}.pdf"'
    
    # Log the export
    AuditLog.objects.create(
        user=request.user,
        company=request.user.company,
        action=AuditLog.ActionType.EXPORT,
        description=f'Exported {queryset.count()} assets to PDF with filters',
        ip_address=request.META.get('REMOTE_ADDR'),
        user_agent=request.META.get('HTTP_USER_AGENT', '')
    )
    
    return response


@login_required
def asset_export_csv(request):
    """Export assets to CSV file."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="assets_export_{datetime.date.today()}.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Asset Number', 'Category', 'Brand', 'Model', 'Serial Number',
        'Status', 'Location', 'Assigned To', 'Purchase Date', 'Purchase Price',
        'Warranty End Date', 'Created At'
    ])
    
    assets = _get_accessible_hardware_assets(request.user).select_related(
        'category', 'brand', 'model', 'assigned_to', 'location'
    )
    
    for asset in assets:
        writer.writerow([
            asset.asset_number,
            asset.category.name if asset.category else '',
            asset.brand.name if asset.brand else '',
            asset.model.name if asset.model else '',
            asset.serial_number,
            asset.get_status_display(),
            asset.get_location_display_abbr(),
            asset.assigned_to.get_display_name() if asset.assigned_to else '',
            asset.purchase_date.strftime('%Y-%m-%d') if asset.purchase_date else '',
            asset.purchase_price,
            asset.warranty_end_date.strftime('%Y-%m-%d') if asset.warranty_end_date else '',
            asset.created_at.strftime('%Y-%m-%d %H:%M:%S')
        ])
    
    # Log the export
    AuditLog.objects.create(
        user=request.user,
        company=request.user.company,
        action=AuditLog.ActionType.EXPORT,
        description=f'Exported {assets.count()} assets to CSV',
        ip_address=request.META.get('REMOTE_ADDR'),
        user_agent=request.META.get('HTTP_USER_AGENT', '')
    )
    
    return response


@login_required
def asset_stats_api(request):
    """API endpoint for asset statistics (for dashboard charts)."""
    accessible_assets = _get_accessible_hardware_assets(request.user)
    
    # Status distribution
    status_stats = {}
    for status, label in Asset.AssetStatus.choices:
        count = accessible_assets.filter(status=status).count()
        status_stats[status] = {
            'label': label,
            'count': count
        }
    
    # Category distribution
    category_stats = list(
        accessible_assets
        .values('category__name')
        .annotate(count=Count('id'))
        .order_by('-count')[:10]
    )
    
    # Location distribution
    location_stats = list(
        accessible_assets
        .values('current_location')
        .annotate(count=Count('id'))
        .order_by('-count')[:10]
    )
    
    # Monthly acquisition trend (last 12 months)
    from django.utils import timezone
    from dateutil.relativedelta import relativedelta
    
    monthly_stats = []
    current_date = timezone.now().date()
    
    for i in range(12):
        month_start = current_date - relativedelta(months=i)
        month_end = month_start + relativedelta(months=1) - relativedelta(days=1)
        
        count = accessible_assets.filter(
            created_at__date__gte=month_start,
            created_at__date__lte=month_end
        ).count()
        
        monthly_stats.insert(0, {
            'month': month_start.strftime('%Y-%m'),
            'count': count
        })
    
    return JsonResponse({
        'status_stats': status_stats,
        'category_stats': category_stats,
        'location_stats': location_stats,
        'monthly_stats': monthly_stats
    })


@login_required
def asset_import_view(request):
    """
    Import assets from CSV or Excel files.
    Supports validation, preview, and bulk creation with error handling.
    """
    rollback_url = _assets_latest_rollback_url(request)

    if request.method == 'POST':
        form = AssetImportForm(request.POST, request.FILES, user=request.user)
        
        if form.is_valid():
            try:
                # Process the uploaded file
                import_run = None
                if not form.cleaned_data['validate_only']:
                    import_run = start_import_run(request.user, ASSETS_IMPORT_MODULE, ASSET_IMPORT_TYPE, total_rows=0)

                result = process_asset_import(
                    file=form.cleaned_data['file'],
                    company=form.cleaned_data['company'],
                    asset_number_mode=form.cleaned_data['asset_number_mode'],
                    asset_number_prefix=form.cleaned_data.get('asset_number_prefix', ''),
                    duplicate_handling=form.cleaned_data['duplicate_handling'],
                    validate_only=form.cleaned_data['validate_only'],
                    user=request.user,
                    import_run=import_run,
                )
                
                if form.cleaned_data['validate_only']:
                    # Preview mode - show validation results
                    messages.info(request, _('File validation completed. Review the results below.'))
                    return render(request, 'assets/import_preview.html', {
                        'form': form,
                        'result': result,
                        'title': _('Asset Import Preview'),
                        'rollback_url': rollback_url,
                    })
                else:
                    # Actual import
                    if result['success']:
                        return render(request, 'common/import_result.html', {
                            'title': _('Asset Import Result'),
                            'total_rows': result.get('total_rows', 0),
                            'processed_rows': result.get('processed_rows', 0),
                            'created': result.get('imported_count', 0),
                            'updated': 0,
                            'skipped': 0,
                            'errors': result.get('errors', [])[:100],
                            'rollback_url': reverse('assets:asset_import_rollback') if import_run and import_run.can_rollback else None,
                            'back_url': reverse('assets:asset_import'),
                        })
                    else:
                        messages.error(request, _('Import failed. Please check the errors below.'))
                        
            except Exception as e:
                messages.error(request, _('Import failed: {error}').format(error=str(e)))
                
        else:
            messages.error(request, _('Please correct the errors below.'))
    else:
        form = AssetImportForm(user=request.user)
    
    return render(request, 'assets/import_form.html', {
        'form': form,
        'title': _('Import Assets'),
        'sample_csv_url': reverse('assets:sample_csv'),
        'rollback_url': rollback_url,
    })


def _normalize_import_header(raw_header):
    header = (str(raw_header or '').strip().lower())
    for marker in ('[required]', '[optional]', '(required)', '(optional)'):
        header = header.replace(marker, '')
    header = header.strip()
    header = header.replace(' ', '_')

    aliases = {
        'model_name': 'model',
        'location_name': 'location',
        'asset_tag': 'asset_number',
    }
    return aliases.get(header, header)


def _normalize_import_row(raw_row):
    normalized = {}
    for key, value in raw_row.items():
        normalized_key = _normalize_import_header(key)
        if not normalized_key:
            continue
        normalized[normalized_key] = value
    return normalized


def _clean_import_value(value):
    if value is None:
        return ''
    if pd.isna(value):
        return ''
    return str(value).strip()


def process_asset_import(file, company, asset_number_mode, asset_number_prefix, 
                        duplicate_handling, validate_only, user, import_run=None):
    """
    Process asset import from CSV or Excel file.
    Returns a dictionary with import results and any errors.
    """
    result = {
        'success': False,
        'imported_count': 0,
        'total_rows': 0,
        'processed_rows': 0,
        'errors': [],
        'warnings': [],
        'processed_assets': []
    }
    
    try:
        # Read file based on extension
        file_extension = file.name.lower().split('.')[-1]
        
        if file_extension == 'csv':
            # Read CSV file
            try:
                csv_reader, _encoding = read_csv_rows_with_fallback(file)
            except UnicodeDecodeError as exc:
                raise ValueError(_('Unable to decode CSV file. Please save it as UTF-8, GBK/GB18030, or Big5.')) from exc
            data_rows = list(csv_reader)
        else:
            # Read Excel file
            df = pd.read_excel(file)
            data_rows = df.to_dict('records')

        data_rows = [_normalize_import_row(row) for row in data_rows]
        result['total_rows'] = len(data_rows)
        if import_run is not None:
            import_run.total_rows = len(data_rows)
            import_run.save(update_fields=['total_rows'])
        
        # Define required and optional columns
        required_columns = ['category', 'brand']
        
        # Validate file has required columns
        if not data_rows:
            result['errors'].append(_('File is empty or has no data rows.'))
            return result
        
        first_row_keys = list(data_rows[0].keys())
        missing_required = [col for col in required_columns if col not in first_row_keys]
        
        if missing_required:
            result['errors'].append(
                _('Missing required columns: {columns}').format(
                    columns=', '.join(missing_required)
                )
            )
            return result
        
        # Process each row
        imported_count = 0
        change_sequence = 0
        
        with transaction.atomic() if not validate_only else transaction.atomic():
            for row_num, row_data in enumerate(data_rows, start=2):  # Start at 2 for Excel row numbers
                try:
                    asset_data = process_asset_row(
                        row_data, row_num, company, asset_number_mode, 
                        asset_number_prefix, duplicate_handling, user
                    )
                    
                    if asset_data['success']:
                        if not validate_only:
                            # Create the asset
                            asset = Asset.objects.create(**asset_data['asset_fields'])
                            imported_count += 1
                            if import_run is not None:
                                change_sequence += 1
                                record_import_change(
                                    import_run,
                                    sequence=change_sequence,
                                    operation=ImportRunChange.ChangeOperation.CREATE,
                                    instance=asset,
                                    row_number=row_num,
                                    after_data=snapshot_instance(asset),
                                )
                            
                            # Log the import action
                            AuditLog.objects.create(
                                company=company,
                                content_object=asset,
                                action=AuditLog.ActionType.IMPORT,
                                user=user,
                                description=f'Imported asset {asset.asset_number} from row {row_num}',
                                metadata={'imported': True, 'row_number': row_num},
                            )
                        else:
                            imported_count += 1
                        
                        result['processed_assets'].append({
                            'row_number': row_num,
                            'status': 'success',
                            'asset_number': asset_data['asset_fields'].get('asset_number', 'Auto-generated')
                        })
                    else:
                        result['errors'].extend([
                            f"Row {row_num}: {error}" for error in asset_data['errors']
                        ])
                        result['processed_assets'].append({
                            'row_number': row_num,
                            'status': 'error',
                            'errors': asset_data['errors']
                        })
                        
                except Exception as e:
                    error_msg = f"Row {row_num}: Unexpected error - {str(e)}"
                    result['errors'].append(error_msg)
                    result['processed_assets'].append({
                        'row_number': row_num,
                        'status': 'error',
                        'errors': [str(e)]
                    })
        
        result['imported_count'] = imported_count
        result['processed_rows'] = len(result['processed_assets'])
        result['success'] = imported_count > 0 or validate_only

        if not validate_only and imported_count > 0:
            AuditLog.objects.create(
                user=user,
                company=company,
                action=AuditLog.ActionType.IMPORT,
                description=f'Imported {imported_count} assets from {file.name}',
                metadata={
                    'imported_count': imported_count,
                    'total_rows': result['total_rows'],
                    'processed_rows': result['processed_rows'],
                    'error_count': len(result['errors']),
                },
            )

        if import_run is not None:
            finalize_import_run(
                import_run,
                created=imported_count,
                updated=0,
                skipped=0,
                error_count=len(result['errors']),
            )
        
    except Exception as e:
        result['errors'].append(f"File processing error: {str(e)}")
        result['processed_rows'] = len(result['processed_assets'])
        if import_run is not None:
            finalize_import_run(
                import_run,
                created=result.get('imported_count', 0),
                updated=0,
                skipped=0,
                error_count=len(result['errors']),
                notes='File processing failed.',
            )
    
    return result


@login_required
@require_POST
def asset_import_rollback_view(request):
    return _perform_assets_rollback(request)


def process_asset_row(row_data, row_num, company, asset_number_mode, 
                     asset_number_prefix, duplicate_handling, user):
    """
    Process a single row of asset data from import file.
    Returns processed asset data or errors.
    """
    result = {
        'success': False,
        'asset_fields': {},
        'errors': [],
        'warnings': [],
    }
    
    try:
        # Basic asset fields
        asset_fields = {
            'company': company,
            'created_by': user,
        }
        
        # Handle asset number (optional - will be auto-generated if not provided)
        if 'asset_number' in row_data and row_data['asset_number']:
            asset_number = _clean_import_value(row_data['asset_number'])
            if asset_number:
                # Apply prefix or mode logic
                if asset_number_mode == 'prefix' and asset_number_prefix:
                    asset_fields['asset_number'] = f"{asset_number_prefix}{asset_number}"
                elif asset_number_mode == 'from_file':
                    asset_fields['asset_number'] = asset_number
                # For 'auto' mode, we leave asset_number empty for auto-generation
        
        # Optional fields with validation
        if 'description' in row_data and row_data['description']:
            asset_fields['description'] = _clean_import_value(row_data['description'])
        
        if 'serial_number' in row_data and row_data['serial_number']:
            serial_number = _clean_import_value(row_data['serial_number'])
            
            # Check for duplicate serial numbers
            if duplicate_handling == 'skip':
                existing_asset = Asset.objects.filter(serial_number=serial_number).first()
                if existing_asset:
                    result['errors'].append(
                        _('Asset with serial number {sn} already exists').format(sn=serial_number)
                    )
                    return result
            
            asset_fields['serial_number'] = serial_number
        
        # Handle category (required)
        category_name = _clean_import_value(row_data.get('category'))
        if not category_name:
            result['errors'].append(_('Category is required.'))
        else:
            try:
                category = hardware_category_queryset().get(name__iexact=category_name)
                asset_fields['category'] = category
            except AssetCategory.DoesNotExist:
                result['errors'].append(
                    _('Category "{category}" not found').format(category=category_name)
                )
        
        # Handle brand (required)
        brand_name = _clean_import_value(row_data.get('brand'))
        if not brand_name:
            result['errors'].append(_('Brand is required.'))
        else:
            try:
                brand = AssetBrand.objects.get(name__iexact=brand_name)
                asset_fields['brand'] = brand
            except AssetBrand.DoesNotExist:
                result['errors'].append(
                    _('Brand "{brand}" not found').format(brand=brand_name)
                )

        # Handle model foreign key (optional)
        model_name = _clean_import_value(row_data.get('model'))
        if model_name:
            if 'brand' not in asset_fields or 'category' not in asset_fields:
                result['errors'].append(_('Model requires valid category and brand.'))
            else:
                model_qs = hardware_model_queryset().filter(
                    name__iexact=model_name,
                    brand=asset_fields['brand'],
                )
                model_qs = model_qs.filter(
                    Q(category=asset_fields['category']) | Q(category__isnull=True)
                )

                if not model_qs.exists():
                    result['errors'].append(
                        _('Model "{model}" not found for brand "{brand}" and category "{category}".').format(
                            model=model_name,
                            brand=asset_fields['brand'].name,
                            category=asset_fields['category'].name,
                        )
                    )
                elif model_qs.count() > 1:
                    result['errors'].append(
                        _('Model "{model}" is ambiguous for brand "{brand}".').format(
                            model=model_name,
                            brand=asset_fields['brand'].name,
                        )
                    )
                else:
                    model = model_qs.first()
                    asset_fields['model'] = model
                    if model and model.category_id is None:
                        result['warnings'].append(
                            _('Model "{model}" has no category assigned in admin.').format(model=model.name)
                        )

        # Handle barcode (optional)
        barcode = _clean_import_value(row_data.get('barcode'))
        if barcode:
            if duplicate_handling == 'skip' and Asset.objects.filter(barcode=barcode).exists():
                result['errors'].append(
                    _('Asset with barcode {barcode} already exists').format(barcode=barcode)
                )
                return result
            asset_fields['barcode'] = barcode
        
        # Handle financial fields
        purchase_price_raw = _clean_import_value(row_data.get('purchase_price'))
        if purchase_price_raw:
            try:
                price = Decimal(purchase_price_raw.replace(',', ''))
                asset_fields['purchase_price'] = price
            except (InvalidOperation, ValueError):
                result['errors'].append(
                    _('Invalid purchase price: {price}').format(price=purchase_price_raw)
                )
        
        # Handle dates
        for date_field in ['purchase_date', 'warranty_end_date']:
            date_raw = _clean_import_value(row_data.get(date_field))
            if date_raw:
                try:
                    date_value = pd.to_datetime(date_raw).date()
                    asset_fields[date_field] = date_value
                except:
                    result['errors'].append(
                        _('Invalid date format for {field}: {value}').format(
                            field=date_field, value=date_raw
                        )
                    )
        
        # Handle other text fields
        for field in ['notes', 'location_zone', 'location_rack', 'location_shelf']:
            if field in row_data and row_data[field]:
                asset_fields[field] = _clean_import_value(row_data[field])

        # Handle location lookup (optional)
        location_name = _clean_import_value(row_data.get('location'))
        if location_name:
            location_qs = Location.objects.filter(company=company, name__iexact=location_name)
            if location_qs.count() == 1:
                location = location_qs.first()
                asset_fields['location'] = location
                asset_fields['current_location'] = location.name
            elif location_qs.count() > 1:
                result['errors'].append(
                    _('Location "{location}" is ambiguous in company "{company}".').format(
                        location=location_name,
                        company=company.name,
                    )
                )
            else:
                result['errors'].append(
                    _('Location "{location}" not found in company "{company}".').format(
                        location=location_name,
                        company=company.name,
                    )
                )

        # Legacy fallback text location
        legacy_location = _clean_import_value(row_data.get('current_location'))
        if legacy_location and 'current_location' not in asset_fields:
            asset_fields['current_location'] = legacy_location
        
        # Handle choice fields
        status_raw = _clean_import_value(row_data.get('status'))
        if status_raw:
            status = status_raw.lower()
            status_choices = dict(Asset.AssetStatus.choices)
            status_key = None
            for key, value in status_choices.items():
                if value.lower() == status or key.lower() == status:
                    status_key = key
                    break
            if status_key:
                asset_fields['status'] = status_key
            else:
                result['errors'].append(
                    _('Unknown status "{status}"').format(status=status_raw)
                )
        
        condition_raw = _clean_import_value(row_data.get('condition'))
        if condition_raw:
            condition = condition_raw.lower()
            condition_choices = dict(Asset.AssetCondition.choices)
            condition_key = None
            for key, value in condition_choices.items():
                if value.lower() == condition or key.lower() == condition:
                    condition_key = key
                    break
            if condition_key:
                asset_fields['condition'] = condition_key
            else:
                result['errors'].append(
                    _('Unknown condition "{condition}"').format(condition=condition_raw)
                )
        
        result['asset_fields'] = asset_fields
        result['success'] = len(result['errors']) == 0
        
    except Exception as e:
        result['errors'].append(f"Processing error: {str(e)}")
    
    return result


@login_required
def download_sample_csv(request):
    """
    Download a sample CSV file for asset import.
    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="asset_import_sample.csv"'
    
    writer = csv.writer(response)
    
    # Write header
    writer.writerow([
        'asset_number [optional]',
        'category [required]',
        'brand [required]',
        'model [optional]',
        'serial_number [optional]',
        'barcode [optional]',
        'description [optional]',
        'purchase_price [optional]',
        'purchase_date [optional]',
        'warranty_end_date [optional]',
        'location_name [optional]',
        'location_zone [optional]',
        'location_rack [optional]',
        'location_shelf [optional]',
        'status [optional]',
        'condition [optional]',
        'notes [optional]',
    ])
    
    # Write sample data
    writer.writerow([
        'LAPTOP-001', 'Laptop', 'Dell', 'XPS 13 9310', 'DL123456789', 'BC-0001',
        'High-performance ultrabook for development work', '1200.00', '2024-01-15',
        '2027-01-15', 'Vanke VMO Warehouse', 'Z1', 'R1', 'S1', 'available', 'good',
        'Includes charger and carrying case'
    ])
    
    writer.writerow([
        '', 'Mobile Device', 'Apple', 'iPhone 14 Pro', 'IP987654321', 'BC-0002',
        'Company mobile phone for executives', '999.00', '2024-02-01',
        '2025-02-01', 'Executive Office', '', '', '', 'assigned', 'excellent',
        'Space Gray, 256GB storage - asset number will be auto-generated'
    ])
    
    writer.writerow([
        'MON-001', 'Monitor', 'HP', 'E27 G5', 'HP555777999', 'BC-0003',
        'External monitor for workstations', '299.99', '2024-01-20',
        '2027-01-20', 'Desk 5A', '', '', '', 'available', 'good',
        '1920x1080 resolution, HDMI and DisplayPort'
    ])
    
    return response


# ============================================================================
# Category Management Views
# ============================================================================

class CategoryListView(LoginRequiredMixin, ListView):
    """List view for asset categories."""
    model = AssetCategory
    template_name = 'assets/category_list.html'
    context_object_name = 'categories'
    paginate_by = 20
    ordering = ['name']

    def get_queryset(self):
        queryset = super().get_queryset().exclude(item_type=AssetCategory.ItemType.SERVICE)
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(description__icontains=search) |
                Q(code__icontains=search)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        return context


class CategoryCreateView(LoginRequiredMixin, CreateView):
    """Create view for asset categories."""
    model = AssetCategory
    form_class = CategoryForm
    template_name = 'assets/category_form.html'
    success_url = reverse_lazy('assets:category_list')

    def form_valid(self, form):
        messages.success(self.request, _('Category created successfully.'))
        return super().form_valid(form)


class CategoryUpdateView(LoginRequiredMixin, UpdateView):
    """Update view for asset categories."""
    model = AssetCategory
    form_class = CategoryForm
    template_name = 'assets/category_form.html'
    success_url = reverse_lazy('assets:category_list')

    def form_valid(self, form):
        messages.success(self.request, _('Category updated successfully.'))
        return super().form_valid(form)


class CategoryDeleteView(LoginRequiredMixin, DeleteView):
    """Delete view for asset categories."""
    model = AssetCategory
    template_name = 'assets/category_confirm_delete.html'
    success_url = reverse_lazy('assets:category_list')

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, _('Category deleted successfully.'))
        return super().delete(request, *args, **kwargs)


# ============================================================================
# Brand Management Views
# ============================================================================

class BrandListView(LoginRequiredMixin, ListView):
    """List view for asset brands."""
    model = AssetBrand
    template_name = 'assets/brand_list.html'
    context_object_name = 'brands'
    paginate_by = 20
    ordering = ['name']

    def get_queryset(self):
        queryset = hardware_brand_queryset()
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(code__icontains=search) |
                Q(description__icontains=search)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        return context


class BrandCreateView(LoginRequiredMixin, CreateView):
    """Create view for asset brands."""
    model = AssetBrand
    form_class = BrandForm
    template_name = 'assets/brand_form.html'
    success_url = reverse_lazy('assets:brand_list')

    def form_valid(self, form):
        messages.success(self.request, _('Brand created successfully.'))
        return super().form_valid(form)


class BrandUpdateView(LoginRequiredMixin, UpdateView):
    """Update view for asset brands."""
    model = AssetBrand
    form_class = BrandForm
    template_name = 'assets/brand_form.html'
    success_url = reverse_lazy('assets:brand_list')

    def form_valid(self, form):
        messages.success(self.request, _('Brand updated successfully.'))
        return super().form_valid(form)


class BrandDeleteView(LoginRequiredMixin, DeleteView):
    """Delete view for asset brands."""
    model = AssetBrand
    template_name = 'assets/brand_confirm_delete.html'
    success_url = reverse_lazy('assets:brand_list')

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, _('Brand deleted successfully.'))
        return super().delete(request, *args, **kwargs)


# ============================================================================
# Model Management Views
# ============================================================================

class ModelListView(LoginRequiredMixin, ListView):
    """List view for asset models."""
    model = AssetModel
    template_name = 'assets/model_list.html'
    context_object_name = 'models'
    paginate_by = 20
    ordering = ['brand__name', 'name']

    def get_queryset(self):
        queryset = hardware_model_queryset().select_related('brand', 'category')
        search = self.request.GET.get('search')
        brand_id = self.request.GET.get('brand')
        unit = self.request.GET.get('unit')
        
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(model_number__icontains=search) |
                Q(brand__name__icontains=search) |
                Q(description__icontains=search)
            )
        
        if brand_id:
            queryset = queryset.filter(brand_id=brand_id)

        if unit:
            queryset = queryset.filter(unit__iexact=unit)
            
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        context['brands'] = hardware_brand_queryset().order_by('name')
        context['selected_brand'] = self.request.GET.get('brand', '')
        context['selected_unit'] = self.request.GET.get('unit', '')
        context['units'] = hardware_model_queryset().exclude(unit='').values_list('unit', flat=True).distinct().order_by('unit')
        return context


class ModelCreateView(LoginRequiredMixin, CreateView):
    """Create view for asset models."""
    model = AssetModel
    form_class = ModelForm
    template_name = 'assets/model_form.html'
    success_url = reverse_lazy('assets:model_list')

    def form_valid(self, form):
        messages.success(self.request, _('Model created successfully.'))
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['brands'] = hardware_brand_queryset().order_by('name')
        return context


class ModelUpdateView(LoginRequiredMixin, UpdateView):
    """Update view for asset models."""
    model = AssetModel
    form_class = ModelForm
    template_name = 'assets/model_form.html'
    success_url = reverse_lazy('assets:model_list')

    def form_valid(self, form):
        messages.success(self.request, _('Model updated successfully.'))
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, _('Please correct the highlighted errors and try again.'))
        return super().form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['brands'] = hardware_brand_queryset().order_by('name')
        return context


class ModelDeleteView(LoginRequiredMixin, DeleteView):
    """Delete view for asset models."""
    model = AssetModel
    template_name = 'assets/model_confirm_delete.html'
    success_url = reverse_lazy('assets:model_list')

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        try:
            response = super().delete(request, *args, **kwargs)
            messages.success(self.request, _('Model deleted successfully.'))
            return response
        except ProtectedError:
            messages.error(
                self.request,
                _(
                    'Cannot delete model "%(model)s" because it is referenced by existing records '
                    '(for example product prices, quotations, or purchase/delivery history). '
                    'Please deactivate or stop using this model instead.'
                ) % {'model': self.object.name},
            )
            return redirect(self.success_url)

    def form_valid(self, form):
        try:
            response = super().form_valid(form)
            messages.success(self.request, _('Model deleted successfully.'))
            return response
        except ProtectedError:
            messages.error(
                self.request,
                _(
                    'Cannot delete model "%(model)s" because it is referenced by existing records '
                    '(for example product prices, quotations, or purchase/delivery history). '
                    'Please deactivate or stop using this model instead.'
                ) % {'model': self.object.name},
            )
            return redirect(self.success_url)


# ============================================================================
# Combined Brands & Models Management View
# ============================================================================

@login_required
def brands_models_view(request):
    """Combined view for managing brands and models."""
    search = request.GET.get('search', '').strip()
    brand_id = request.GET.get('brand', '').strip()
    unit = request.GET.get('unit', '').strip()

    models_qs = hardware_model_queryset().select_related('brand', 'category').order_by('brand__name', 'name')
    if search:
        models_qs = models_qs.filter(
            Q(name__icontains=search)
            | Q(model_number__icontains=search)
            | Q(description__icontains=search)
            | Q(brand__name__icontains=search)
        )
    if brand_id:
        models_qs = models_qs.filter(brand_id=brand_id)
    if unit:
        models_qs = models_qs.filter(unit__iexact=unit)

    brands = hardware_brand_queryset().filter(models__in=models_qs).distinct().order_by('name')
    context = {
        'brands': brands,
        'models': models_qs,
        'search_query': search,
        'selected_brand': brand_id,
        'selected_unit': unit,
        'brand_filter_options': hardware_brand_queryset().order_by('name'),
        'unit_filter_options': hardware_model_queryset().exclude(unit='').values_list('unit', flat=True).distinct().order_by('unit'),
    }
    return render(request, 'assets/brands_models.html', context)

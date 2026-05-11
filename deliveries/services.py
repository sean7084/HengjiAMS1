"""Services for delivery template filling and PDF generation."""

from pathlib import Path
import shutil
import subprocess

from django.conf import settings
from django.db.models import Q
from django.template.loader import render_to_string

from assets.models import Asset

from .models import DeliveryOrder

from openpyxl import load_workbook
from weasyprint import HTML


INTERNAL_WAREHOUSE_LOCATION_ID = 3


def is_service_quotation_item(item):
    if getattr(item, 'service_item_id', None):
        return True
    product_price = getattr(item, 'product_price', None)
    return bool(getattr(product_price, 'service_item_id', None))


def split_quotation_items_for_delivery(quotation):
    hardware_items = []
    service_items = []

    if quotation is None:
        return hardware_items, service_items

    quotation_items = quotation.items.select_related(
        'product_price__brand',
        'product_price__model__category',
        'product_price__service_item',
        'service_item',
    )
    for item in quotation_items:
        if is_service_quotation_item(item):
            service_items.append(item)
        else:
            hardware_items.append(item)

    return hardware_items, service_items


def build_dispatch_asset_assignments(quotation, assets):
    hardware_items, _ = split_quotation_items_for_delivery(quotation)
    available_assets_by_key = {}
    for asset in assets:
        available_assets_by_key.setdefault((asset.brand_id, asset.model_id), []).append(asset)

    assignments = []
    for item in hardware_items:
        if not item.product_price_id or not item.product_price.brand_id or not item.product_price.model_id:
            return None

        key = (item.product_price.brand_id, item.product_price.model_id)
        matching_assets = available_assets_by_key.get(key, [])
        if len(matching_assets) < item.quantity:
            return None

        selected_assets = matching_assets[:item.quantity]
        available_assets_by_key[key] = matching_assets[item.quantity:]
        assignments.append((item, selected_assets))

    return assignments


def create_delivery_items_from_asset_assignments(delivery_order, assignments):
    from .models import DeliveryItem

    for quotation_item, assets in assignments:
        for asset in assets:
            DeliveryItem.objects.create(
                delivery_order=delivery_order,
                asset=asset,
                quotation_item=quotation_item,
                quantity=1,
                user_brand=quotation_item.user_brand,
                user_name=quotation_item.user_name,
            )


def sync_service_delivery_items(delivery_order):
    from .models import DeliveryItem

    _, service_items = split_quotation_items_for_delivery(delivery_order.quotation)
    for quotation_item in service_items:
        DeliveryItem.objects.update_or_create(
            delivery_order=delivery_order,
            asset=None,
            quotation_item=quotation_item,
            defaults={
                'serial_number': '',
                'brand_name': quotation_item.brand_name,
                'product_description': quotation_item.product_description,
                'user_brand': quotation_item.user_brand,
                'user_name': quotation_item.user_name,
                'quantity': quotation_item.quantity,
            },
        )


def get_dispatch_asset_queryset(quotation):
    if quotation is None:
        return Asset.objects.none()

    item_filters = Q()
    hardware_items, _ = split_quotation_items_for_delivery(quotation)
    for item in hardware_items:
        if not item.product_price_id or not item.product_price.brand_id or not item.product_price.model_id:
            continue
        item_filters |= Q(brand_id=item.product_price.brand_id, model_id=item.product_price.model_id)

    if not item_filters:
        return Asset.objects.none()

    active_delivery_statuses = [
        DeliveryOrder.Status.PENDING,
        DeliveryOrder.Status.DISPATCHED,
    ]

    return Asset.objects.filter(
        status=Asset.AssetStatus.AVAILABLE,
        location_id=INTERNAL_WAREHOUSE_LOCATION_ID,
    ).filter(
        item_filters,
    ).exclude(
        delivery_items__delivery_order__status__in=active_delivery_statuses,
    ).select_related('brand', 'model').distinct().order_by('asset_number')


def _find_label_cell(sheet, label):
    label = str(label).strip()
    for row in sheet.iter_rows():
        for cell in row:
            if str(cell.value).strip() == label:
                return cell
    return None


def _fill_by_label(sheet, label, value):
    cell = _find_label_cell(sheet, label)
    if not cell:
        return False

    target = sheet.cell(row=cell.row, column=cell.column + 1)
    target.value = value
    return True


def _find_item_header_map(sheet):
    expected = {
        '序列号': 'serial_number',
        '品牌': 'brand_name',
        '商品描述': 'product_description',
        '采购方品牌': 'user_brand',
        '采购方用户': 'user_name',
        '数量': 'quantity',
    }

    for row in sheet.iter_rows(min_row=1, max_row=150):
        header_map = {}
        for cell in row:
            text = str(cell.value).strip() if cell.value is not None else ''
            if text in expected:
                header_map[expected[text]] = cell.column

        if len(header_map) >= 3:
            return row[0].row, header_map

    return None, {}


def fill_delivery_template(delivery_order):
    """Fill delivery Excel template and return generated xlsx path."""
    template_path = Path(settings.BASE_DIR) / 'template_files' / '签收单 template.xlsx'
    if not template_path.exists():
        raise FileNotFoundError(f'Template not found: {template_path}')

    workbook = load_workbook(template_path)
    sheet = workbook.active

    _fill_by_label(sheet, '订货方', delivery_order.quotation.customer.name)
    _fill_by_label(sheet, '收货人', delivery_order.receiver_name)
    _fill_by_label(sheet, '电话', delivery_order.receiver_phone)
    _fill_by_label(sheet, '交货地址', delivery_order.delivery_address)
    _fill_by_label(sheet, '交货方式', delivery_order.delivery_method)

    # Try common date labels
    if not _fill_by_label(sheet, '日期', delivery_order.delivery_date.strftime('%Y-%m-%d')):
        _fill_by_label(sheet, '送货日期', delivery_order.delivery_date.strftime('%Y-%m-%d'))

    start_row, header_map = _find_item_header_map(sheet)
    if start_row and header_map:
        row_idx = start_row + 1
        for item in delivery_order.items.all().order_by('id'):
            values = {
                'serial_number': item.serial_number,
                'brand_name': item.brand_name,
                'product_description': item.product_description,
                'user_brand': item.user_brand,
                'user_name': item.user_name,
                'quantity': item.quantity,
            }
            for key, col in header_map.items():
                sheet.cell(row=row_idx, column=col).value = values.get(key, '')
            row_idx += 1

    output_dir = Path(settings.MEDIA_ROOT) / 'deliveries' / 'generated'
    output_dir.mkdir(parents=True, exist_ok=True)

    xlsx_path = output_dir / f'{delivery_order.delivery_number}.xlsx'
    workbook.save(xlsx_path)

    return xlsx_path


def _find_soffice_binary():
    candidates = [
        shutil.which('soffice'),
        shutil.which('soffice.exe'),
        r'C:\Program Files\LibreOffice\program\soffice.exe',
        r'C:\Program Files (x86)\LibreOffice\program\soffice.exe',
    ]

    for candidate in candidates:
        if candidate and Path(candidate).exists():
            return str(candidate)

    return None


def convert_xlsx_to_pdf(xlsx_path):
    """Convert xlsx to pdf with LibreOffice headless. Returns pdf path or None."""
    soffice = _find_soffice_binary()
    if not soffice:
        return None

    xlsx_path = Path(xlsx_path)
    outdir = str(xlsx_path.parent)

    try:
        subprocess.run(
            [
                soffice,
                '--headless',
                '--convert-to',
                'pdf',
                '--outdir',
                outdir,
                str(xlsx_path),
            ],
            check=False,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            timeout=120,
        )
    except Exception:
        return None

    pdf_path = xlsx_path.with_suffix('.pdf')
    return pdf_path if pdf_path.exists() else None


def render_delivery_pdf_html(delivery_order):
    """Render delivery order PDF bytes from Django HTML template (LibreOffice-free path)."""
    ordered_items = list(delivery_order.items.all().order_by('id'))
    display_rows = []
    for item in ordered_items:
        display_rows.append({
            'serial_number': item.serial_number,
            'brand_name': item.brand_name,
            'product_description': item.product_description,
            'user_brand': item.user_brand,
            'user_name': item.user_name,
            'quantity': item.quantity,
        })

    while len(display_rows) < 6:
        display_rows.append({
            'serial_number': '',
            'brand_name': '',
            'product_description': '',
            'user_brand': '',
            'user_name': '',
            'quantity': '',
        })

    delivery_method_display = (delivery_order.delivery_method or '').strip() or '-'

    context = {
        'delivery': delivery_order,
        'display_rows': display_rows,
        'delivery_method_display': delivery_method_display,
        'logo_path': (Path(settings.BASE_DIR) / 'static' / 'images' / 'quotation_template_logo.png').resolve().as_uri(),
        'acceptance_text': '订货方或收货人在本签收单或快递公司承运单上签收后，即视为对以上商品进行了签收。',
    }

    html = render_to_string('deliveries/pdf_excel_style.html', context)
    return HTML(string=html, base_url=str(settings.BASE_DIR)).write_pdf()

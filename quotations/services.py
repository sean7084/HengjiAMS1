"""Services for quotation template filling and PDF conversion."""

from pathlib import Path
from decimal import Decimal
import shutil
import subprocess

from django.conf import settings
from django.template.loader import render_to_string

from openpyxl import load_workbook
from weasyprint import HTML

from quotations.template_registry import get_quotation_template_definition


def _find_label_cell(sheet, label):
    label = str(label).strip()
    for row in sheet.iter_rows():
        for cell in row:
            if str(cell.value).strip() == label:
                return cell
    return None


def _fill_by_label(sheet, label, value):
    cell = _find_label_cell(sheet, label)
    if not cell:
        return False

    target = sheet.cell(row=cell.row, column=cell.column + 1)
    target.value = value
    return True


def _find_item_header_map(sheet):
    expected = {
        '序列号': 'serial_number',
        '品牌': 'brand_name',
        '商品描述': 'product_description',
        '采购方品牌': 'user_brand',
        '采购方用户': 'user_name',
        '单位': 'unit',
        '数量': 'quantity',
        '未税单价': 'unit_price',
        '税率': 'tax_rate',
        '未税金额': 'line_total_without_tax',
        '税额': 'tax_amount',
        '含税金额': 'line_total_with_tax',
    }

    for row in sheet.iter_rows(min_row=1, max_row=180):
        header_map = {}
        for cell in row:
            text = str(cell.value).strip() if cell.value is not None else ''
            if text in expected:
                header_map[expected[text]] = cell.column

        if len(header_map) >= 4:
            return row[0].row, header_map

    return None, {}


def fill_quotation_template(quotation):
    """Fill quotation Excel template and return generated xlsx path."""
    template_path = Path(settings.BASE_DIR) / 'template_files' / 'quotation_template.xlsx'
    if not template_path.exists():
        raise FileNotFoundError(f'Template not found: {template_path}')

    workbook = load_workbook(template_path)
    sheet = workbook.active

    _fill_by_label(sheet, '报价单号', quotation.quotation_number)
    _fill_by_label(sheet, '报价日期', quotation.quotation_date.strftime('%Y-%m-%d'))
    _fill_by_label(sheet, '有效期至', quotation.valid_until.strftime('%Y-%m-%d'))
    _fill_by_label(sheet, '客户名称', quotation.customer.name)
    _fill_by_label(sheet, '客户代码', quotation.customer.code)
    _fill_by_label(sheet, '联系人', quotation.attn)
    _fill_by_label(sheet, '电话', quotation.tel)
    _fill_by_label(sheet, '未税合计', quotation.total_without_tax)
    _fill_by_label(sheet, '税额合计', quotation.total_tax)
    _fill_by_label(sheet, '含税合计', quotation.total_with_tax)
    _fill_by_label(sheet, '备注', quotation.get_effective_remarks())

    start_row, header_map = _find_item_header_map(sheet)
    if start_row and header_map:
        row_idx = start_row + 1
        for index, item in enumerate(quotation.items.all().order_by('id'), start=1):
            values = {
                'serial_number': index,
                'brand_name': item.brand_name,
                'product_description': item.product_description,
                'user_brand': item.user_brand,
                'user_name': item.user_name,
                'unit': item.unit,
                'quantity': item.quantity,
                'unit_price': item.unit_price,
                'tax_rate': item.tax_rate,
                'line_total_without_tax': item.line_total_without_tax,
                'tax_amount': item.tax_amount,
                'line_total_with_tax': item.line_total_with_tax,
            }
            for key, col in header_map.items():
                sheet.cell(row=row_idx, column=col).value = values.get(key, '')
            row_idx += 1

    output_dir = Path(settings.MEDIA_ROOT) / 'quotations' / 'generated'
    output_dir.mkdir(parents=True, exist_ok=True)

    xlsx_path = output_dir / f'{quotation.quotation_number}.xlsx'
    workbook.save(xlsx_path)
    return xlsx_path


def _find_soffice_binary():
    candidates = [
        shutil.which('soffice'),
        shutil.which('soffice.exe'),
        r'C:\Program Files\LibreOffice\program\soffice.exe',
        r'C:\Program Files (x86)\LibreOffice\program\soffice.exe',
    ]

    for candidate in candidates:
        if candidate and Path(candidate).exists():
            return str(candidate)

    return None


def convert_xlsx_to_pdf(xlsx_path):
    """Convert xlsx to pdf with LibreOffice headless. Returns pdf path or None."""
    soffice = _find_soffice_binary()
    if not soffice:
        return None

    xlsx_path = Path(xlsx_path)
    outdir = str(xlsx_path.parent)

    try:
        subprocess.run(
            [
                soffice,
                '--headless',
                '--convert-to',
                'pdf',
                '--outdir',
                outdir,
                str(xlsx_path),
            ],
            check=False,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            timeout=120,
        )
    except Exception:
        return None

    pdf_path = xlsx_path.with_suffix('.pdf')
    return pdf_path if pdf_path.exists() else None


def _build_pdf_display_row(item, index):
    unit_price_with_tax = item.unit_price * (Decimal('1.00') + (item.tax_rate / Decimal('100.00')))
    return {
        'index': index,
        'model_number': item.model_number,
        'brand_name': item.brand_name,
        'product_description': item.product_description,
        'user_brand': item.user_brand,
        'user_name': item.user_name,
        'unit': item.unit,
        'unit_price': item.unit_price,
        'unit_price_with_tax': unit_price_with_tax,
        'quantity': item.quantity,
        'line_total_without_tax': item.line_total_without_tax,
        'line_total_with_tax': item.line_total_with_tax,
        'tax_amount': item.tax_amount,
    }


def _build_empty_pdf_display_row(index):
    return {
        'index': index,
        'model_number': '',
        'brand_name': '',
        'product_description': '',
        'user_brand': '',
        'user_name': '',
        'unit': '',
        'unit_price': '',
        'unit_price_with_tax': '',
        'quantity': '',
        'line_total_without_tax': '',
        'line_total_with_tax': '',
        'tax_amount': '',
    }


def _build_pdf_display_rows(items, minimum_rows=1):
    rows = [_build_pdf_display_row(item, index) for index, item in enumerate(items, start=1)]
    while len(rows) < minimum_rows:
        rows.append(_build_empty_pdf_display_row(len(rows) + 1))
    return rows


def _build_pdf_section_totals(items):
    total_without_tax = Decimal('0.00')
    total_tax = Decimal('0.00')
    total_with_tax = Decimal('0.00')

    for item in items:
        total_without_tax += item.line_total_without_tax
        total_tax += item.tax_amount
        total_with_tax += item.line_total_with_tax

    vat_percent = Decimal('0.00')
    if total_without_tax:
        vat_percent = (total_tax / total_without_tax) * Decimal('100.00')

    return {
        'total_without_tax': total_without_tax,
        'total_tax': total_tax,
        'total_with_tax': total_with_tax,
        'vat_percent': vat_percent,
    }


def split_quotation_items_for_pdf(ordered_items):
    hardware_items = [item for item in ordered_items if not getattr(item, 'service_item_id', None)]
    service_items = [item for item in ordered_items if getattr(item, 'service_item_id', None)]
    return {
        'hardware_items': hardware_items,
        'service_items': service_items,
        'hardware_display_rows': _build_pdf_display_rows(hardware_items),
        'service_display_rows': _build_pdf_display_rows(service_items),
        'hardware_totals': _build_pdf_section_totals(hardware_items),
        'service_totals': _build_pdf_section_totals(service_items),
    }


def resolve_quotation_pdf_template(quotation, template_code=None):
    return get_quotation_template_definition(template_code or quotation.pdf_template)


def render_quotation_pdf_html(quotation, template_code=None):
    """Render quotation PDF bytes from Django HTML template (LibreOffice-free path)."""
    ordered_items = list(quotation.items.all().order_by('id'))
    pdf_sections = split_quotation_items_for_pdf(ordered_items)
    template_definition = resolve_quotation_pdf_template(quotation, template_code=template_code)

    vat_percent = Decimal('0.00')
    if quotation.total_without_tax:
        vat_percent = (quotation.total_tax / quotation.total_without_tax) * Decimal('100')

    context = {
        'quotation': quotation,
        'items': ordered_items,
        'vat_percent': vat_percent,
        'template_definition': template_definition,
        'remarks_text': quotation.get_effective_remarks(ordered_items=ordered_items),
        'logo_path': (Path(settings.BASE_DIR) / 'static' / 'images' / 'quotation_template_logo.png').resolve().as_uri(),
        'prepared_by_company': '上海珩际信息科技有限公司',
    }
    context.update(pdf_sections)
    html = render_to_string(template_definition['template_path'], context)
    return HTML(string=html, base_url=str(settings.BASE_DIR)).write_pdf()

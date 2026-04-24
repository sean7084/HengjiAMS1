from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP
from email.message import EmailMessage as SMTPEmailMessage
from email.utils import formataddr
from pathlib import Path
import shutil
import subprocess
import mimetypes
import smtplib

import openpyxl
from openpyxl import load_workbook

from django.conf import settings
from django.core.mail import EmailMessage as DjangoEmailMessage
from django.db import transaction
from django.http import HttpResponse
from django.utils import timezone

from accounts.mailbox_sync import cache_dispatch_outbox_message
from accounts.models import UserMailboxSettings

from .models import EmailDispatch, InvoiceInfo, InvoiceInfoItem, WeeklyOrderBatch


HEADER_ALIASES = {
    'kering_group_po_number': [
        'kering group po number',
        'kering po number',
        'po number',
        'po no',
        'po',
    ],
    'internal_order': [
        'internal order',
        'io',
        'internal order number',
        'internal order no',
    ],
    'sap_cost_center': [
        'sap cost center',
        'cost center',
        'sap cc',
    ],
}


@dataclass
class ImportResult:
    total_rows: int
    created_rows: int


def _normalize_header(value):
    return ' '.join(str(value or '').strip().lower().replace('_', ' ').split())


def _resolve_header_map(sheet):
    header_row = [cell.value for cell in sheet[1]]
    normalized = {
        _normalize_header(value): index
        for index, value in enumerate(header_row)
        if value is not None
    }

    resolved = {}
    for canonical, aliases in HEADER_ALIASES.items():
        candidates = [_normalize_header(canonical)] + [_normalize_header(alias) for alias in aliases]
        match = next((name for name in candidates if name in normalized), None)
        if not match:
            raise ValueError(
                f"Missing required column for '{canonical}'. Accepted headers: {', '.join(aliases)}"
            )
        resolved[canonical] = normalized[match]

    return resolved


def _clean_text(value):
    return str(value or '').strip()


def process_sharepoint_batch(batch, user=None):
    workbook = openpyxl.load_workbook(batch.sharepoint_file, data_only=True)
    sheet = workbook.active
    header_map = _resolve_header_map(sheet)

    process_date = timezone.localdate()
    parsed_rows = []
    seen_keys = set()

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row_values = list(row)
        if not any(row_values):
            continue

        po_number = _clean_text(row_values[header_map['kering_group_po_number']])
        internal_order = _clean_text(row_values[header_map['internal_order']])
        sap_cost_center = _clean_text(row_values[header_map['sap_cost_center']])

        if not po_number or not internal_order or not sap_cost_center:
            raise ValueError(f'Row {row_number}: required fields cannot be empty.')

        duplicate_key = (po_number, internal_order, sap_cost_center)
        if duplicate_key in seen_keys:
            raise ValueError(
                f'Row {row_number}: duplicate entry found in file for PO/Internal Order/SAP Cost Center.'
            )

        exists = InvoiceInfo.objects.filter(
            kering_group_po_number=po_number,
            internal_order=internal_order,
            sap_cost_center=sap_cost_center,
        ).exists()
        if exists:
            raise ValueError(
                f'Row {row_number}: duplicate entry already exists in system for PO/Internal Order/SAP Cost Center.'
            )

        seen_keys.add(duplicate_key)
        parsed_rows.append(
            {
                'row_number': row_number,
                'kering_group_po_number': po_number,
                'internal_order': internal_order,
                'sap_cost_center': sap_cost_center,
                'invoice_date': process_date,
            }
        )

    with transaction.atomic():
        batch.status = WeeklyOrderBatch.Status.PROCESSING
        batch.processed_by = user
        batch.failure_reason = ''
        batch.failed_row_number = None
        batch.save(update_fields=['status', 'processed_by', 'failure_reason', 'failed_row_number'])

        created = 0
        for row_data in parsed_rows:
            InvoiceInfo.objects.create(
                weekly_batch=batch,
                source_row_number=row_data['row_number'],
                kering_group_po_number=row_data['kering_group_po_number'],
                internal_order=row_data['internal_order'],
                sap_cost_center=row_data['sap_cost_center'],
                invoice_date=row_data['invoice_date'],
            )
            created += 1

        batch.total_rows = len(parsed_rows)
        batch.created_rows = created
        batch.status = WeeklyOrderBatch.Status.PROCESSED
        batch.processed_at = timezone.now()
        batch.save(update_fields=['total_rows', 'created_rows', 'status', 'processed_at'])

    return ImportResult(total_rows=len(parsed_rows), created_rows=created)


def _q2(value):
    return Decimal(value).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)


def recalculate_invoice_from_delivery(invoice):
    """Rebuild invoice line items and totals from linked delivery items."""
    delivery = invoice.delivery_order

    if not delivery:
        return

    quotation = invoice.quotation or delivery.quotation

    with transaction.atomic():
        invoice.items.all().delete()

        total_net = Decimal('0.00')
        total_tax = Decimal('0.00')

        for index, d_item in enumerate(delivery.items.select_related('asset', 'asset__brand', 'asset__model'), start=1):
            q_item = None
            if quotation and d_item.asset_id:
                q_item = quotation.items.filter(
                    product_price__brand=d_item.asset.brand,
                    product_price__model=d_item.asset.model,
                ).first()

            quantity = Decimal(str(d_item.quantity or 1))
            unit_price = q_item.unit_price if q_item else Decimal('0.00')
            tax_rate = q_item.tax_rate if q_item else invoice.tax_rate

            line_net = _q2(unit_price * quantity)
            line_tax = _q2(line_net * (tax_rate / Decimal('100.00')))
            line_gross = _q2(line_net + line_tax)

            description = d_item.product_description or (q_item.product_description if q_item else 'Delivery Item')

            InvoiceInfoItem.objects.create(
                invoice_info=invoice,
                line_number=index,
                description=description,
                unit_price=_q2(unit_price),
                quantity=quantity,
                total_price=line_gross,
                net_amount=line_net,
                tax_amount=line_tax,
                gross_amount=line_gross,
                tax_rate=_q2(tax_rate),
            )

            total_net += line_net
            total_tax += line_tax

        total_gross = _q2(total_net + total_tax)
        effective_tax_rate = Decimal('0.00')
        if total_net > 0:
            effective_tax_rate = _q2((total_tax / total_net) * Decimal('100.00'))

        invoice.quotation = quotation
        invoice.bill_to = quotation.customer.name if quotation else invoice.bill_to
        invoice.net_amount = _q2(total_net)
        invoice.tax_amount = _q2(total_tax)
        invoice.gross_amount = total_gross
        invoice.total_amount = total_gross
        invoice.tax_rate = effective_tax_rate
        invoice.save(
            update_fields=[
                'quotation',
                'bill_to',
                'net_amount',
                'tax_amount',
                'gross_amount',
                'total_amount',
                'tax_rate',
                'updated_at',
            ]
        )


def _find_label_cell(sheet, label):
    label = str(label).strip()
    for row in sheet.iter_rows():
        for cell in row:
            if str(cell.value).strip() == label:
                return cell
    return None


def _fill_by_label(sheet, labels, value):
    if isinstance(labels, str):
        labels = [labels]

    for label in labels:
        cell = _find_label_cell(sheet, label)
        if cell:
            target = sheet.cell(row=cell.row, column=cell.column + 1)
            target.value = value
            return True

    return False


def _find_invoice_item_header(sheet):
    expected = {
        'description': ['Description', '商品描述', '描述'],
        'unit_price': ['Unit Price', '单价'],
        'quantity': ['Quantity', '数量'],
        'net_amount': ['Net Amount', '未税金额', 'Net'],
        'tax_amount': ['Tax Amount', '税额', 'Tax'],
        'gross_amount': ['Gross Amount', '含税金额', 'Total'],
    }

    for row in sheet.iter_rows(min_row=1, max_row=200):
        mapped = {}
        for cell in row:
            text = str(cell.value).strip() if cell.value is not None else ''
            for key, candidates in expected.items():
                if text in candidates:
                    mapped[key] = cell.column

        if len(mapped) >= 3:
            return row[0].row, mapped

    return None, {}


def fill_invoice_template(invoice):
    template_path = Path(settings.BASE_DIR) / 'template_files' / 'invoice information template.xlsx'
    if not template_path.exists():
        raise FileNotFoundError(f'Template not found: {template_path}')

    workbook = load_workbook(template_path)
    sheet = workbook.active

    _fill_by_label(sheet, ['Bill To', '订货方'], invoice.bill_to)
    _fill_by_label(sheet, ['PI Number', 'Invoice Number', '发票号'], invoice.invoice_number)
    _fill_by_label(sheet, ['Invoice Date', '开票日期'], invoice.invoice_date.strftime('%Y-%m-%d'))
    _fill_by_label(
        sheet,
        ['Due Date', 'Payment Due Date', '付款到期日'],
        invoice.payment_due_date.strftime('%Y-%m-%d') if invoice.payment_due_date else '',
    )
    _fill_by_label(sheet, ['PO Number', 'Kering Group PO Number'], invoice.kering_group_po_number)
    _fill_by_label(sheet, ['Internal Order', 'IO'], invoice.internal_order)
    _fill_by_label(sheet, ['SAP Cost Center', 'Cost Center'], invoice.sap_cost_center)
    _fill_by_label(sheet, ['Tax Rate', '税率'], float(invoice.tax_rate))
    _fill_by_label(sheet, ['Net Amount', '未税金额'], float(invoice.net_amount))
    _fill_by_label(sheet, ['Tax Amount', '税额'], float(invoice.tax_amount))
    _fill_by_label(sheet, ['Gross Amount', '含税金额', 'Total Amount'], float(invoice.gross_amount))

    start_row, header_map = _find_invoice_item_header(sheet)
    if start_row and header_map:
        row_idx = start_row + 1
        for item in invoice.items.all().order_by('line_number'):
            row_values = {
                'description': item.description,
                'unit_price': float(item.unit_price),
                'quantity': float(item.quantity),
                'net_amount': float(item.net_amount),
                'tax_amount': float(item.tax_amount),
                'gross_amount': float(item.gross_amount),
            }
            for key, col in header_map.items():
                sheet.cell(row=row_idx, column=col).value = row_values.get(key, '')
            row_idx += 1

    output_dir = Path(settings.MEDIA_ROOT) / 'invoices' / 'generated'
    output_dir.mkdir(parents=True, exist_ok=True)

    xlsx_path = output_dir / f'{invoice.invoice_number}.xlsx'
    workbook.save(xlsx_path)
    return xlsx_path


def _find_soffice_binary():
    candidates = [
        shutil.which('soffice'),
        shutil.which('soffice.exe'),
        r'C:\Program Files\LibreOffice\program\soffice.exe',
        r'C:\Program Files (x86)\LibreOffice\program\soffice.exe',
    ]

    for candidate in candidates:
        if candidate and Path(candidate).exists():
            return str(candidate)

    return None


def convert_xlsx_to_pdf(xlsx_path):
    soffice = _find_soffice_binary()
    if not soffice:
        return None

    xlsx_path = Path(xlsx_path)
    outdir = str(xlsx_path.parent)

    try:
        subprocess.run(
            [
                soffice,
                '--headless',
                '--convert-to',
                'pdf',
                '--outdir',
                outdir,
                str(xlsx_path),
            ],
            check=False,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            timeout=120,
        )
    except Exception:
        return None

    pdf_path = xlsx_path.with_suffix('.pdf')
    return pdf_path if pdf_path.exists() else None


def export_invoice_infos_to_excel(queryset):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Invoice Info'

    headers = [
        'Invoice Number',
        'Invoice Date',
        'Bill To',
        'PO Number',
        'Internal Order',
        'SAP Cost Center',
        'Net Amount',
        'Tax Amount',
        'Gross Amount',
        'Tax Rate',
    ]
    sheet.append(headers)

    for invoice in queryset:
        sheet.append(
            [
                invoice.invoice_number,
                invoice.invoice_date.strftime('%Y-%m-%d') if invoice.invoice_date else '',
                invoice.bill_to,
                invoice.kering_group_po_number,
                invoice.internal_order,
                invoice.sap_cost_center,
                float(invoice.net_amount),
                float(invoice.tax_amount),
                float(invoice.gross_amount),
                float(invoice.tax_rate),
            ]
        )

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=invoice_info_export.xlsx'
    workbook.save(response)
    return response


def _parse_recipients(value):
    return [item.strip() for item in (value or '').split(',') if item.strip()]


def collect_email_attachments(dispatch):
    """Collect all related quotation, delivery, and invoice files for email dispatch."""
    files = []

    quotation = dispatch.quotation
    for attachment in quotation.attachments.all().order_by('uploaded_at'):
        if not attachment.file:
            continue
        file_path = Path(attachment.file.path)
        if not file_path.exists():
            continue
        files.append(
            {
                'category': 'quotation_attachment',
                'label': attachment.get_attachment_type_display(),
                'name': file_path.name,
                'path': str(file_path),
            }
        )

    if dispatch.delivery_order_id:
        delivery = dispatch.delivery_order
        if delivery.signed_file:
            signed_path = Path(delivery.signed_file.path)
            if signed_path.exists():
                files.append(
                    {
                        'category': 'delivery_signed',
                        'label': 'Signed Delivery Copy',
                        'name': signed_path.name,
                        'path': str(signed_path),
                    }
                )

        from deliveries.services import convert_xlsx_to_pdf as delivery_convert_xlsx_to_pdf
        from deliveries.services import fill_delivery_template

        try:
            delivery_xlsx = fill_delivery_template(delivery)
            files.append(
                {
                    'category': 'delivery_document',
                    'label': 'Delivery Document (Excel)',
                    'name': delivery_xlsx.name,
                    'path': str(delivery_xlsx),
                }
            )

            delivery_pdf = delivery_convert_xlsx_to_pdf(delivery_xlsx)
            if delivery_pdf:
                files.append(
                    {
                        'category': 'delivery_document',
                        'label': 'Delivery Document (PDF)',
                        'name': delivery_pdf.name,
                        'path': str(delivery_pdf),
                    }
                )
        except FileNotFoundError:
            pass

    if dispatch.invoice_info_id:
        invoice = dispatch.invoice_info
        invoice_xlsx = fill_invoice_template(invoice)
        files.append(
            {
                'category': 'invoice_document',
                'label': 'Invoice Information (Excel)',
                'name': invoice_xlsx.name,
                'path': str(invoice_xlsx),
            }
        )
        invoice_pdf = convert_xlsx_to_pdf(invoice_xlsx)
        if invoice_pdf:
            files.append(
                {
                    'category': 'invoice_document',
                    'label': 'Invoice Information (PDF)',
                    'name': invoice_pdf.name,
                    'path': str(invoice_pdf),
                }
            )

    dedup = {}
    for entry in files:
        dedup[entry['path']] = entry

    return list(dedup.values())


def _build_smtp_message(dispatch, attachments, mailbox_settings):
    message = SMTPEmailMessage()
    message['Subject'] = dispatch.subject
    message['From'] = formataddr((mailbox_settings.display_name or mailbox_settings.user.get_display_name(), mailbox_settings.email_address))
    message['To'] = ', '.join(_parse_recipients(dispatch.sent_to))
    if dispatch.cc:
        message['Cc'] = ', '.join(_parse_recipients(dispatch.cc))
    if dispatch.bcc:
        message['Bcc'] = ', '.join(_parse_recipients(dispatch.bcc))
    message.set_content(dispatch.body or '')

    for entry in attachments:
        path = Path(entry['path'])
        if not path.exists():
            continue
        mime_type, _ = mimetypes.guess_type(path.name)
        main_type, sub_type = (mime_type or 'application/octet-stream').split('/', 1)
        with open(path, 'rb') as file_obj:
            message.add_attachment(file_obj.read(), maintype=main_type, subtype=sub_type, filename=path.name)

    return message


def _send_via_user_mailbox(dispatch, attachments):
    creator = dispatch.created_by
    mailbox_settings = getattr(creator, 'mailbox_settings', None) if creator else None
    if not mailbox_settings or not mailbox_settings.is_active or not mailbox_settings.password:
        return False

    message = _build_smtp_message(dispatch, attachments, mailbox_settings)
    recipients = _parse_recipients(dispatch.sent_to) + _parse_recipients(dispatch.cc) + _parse_recipients(dispatch.bcc)

    if mailbox_settings.smtp_security == UserMailboxSettings.ConnectionSecurity.SSL_TLS:
        client = smtplib.SMTP_SSL(mailbox_settings.smtp_host, mailbox_settings.smtp_port)
    else:
        client = smtplib.SMTP(mailbox_settings.smtp_host, mailbox_settings.smtp_port)
        if mailbox_settings.smtp_security == UserMailboxSettings.ConnectionSecurity.STARTTLS:
            client.starttls()

    try:
        client.login(mailbox_settings.username, mailbox_settings.password)
        client.send_message(message, to_addrs=recipients)
    except Exception as exc:
        mailbox_settings.last_connection_test_at = timezone.now()
        mailbox_settings.last_connection_status = 'failed'
        mailbox_settings.last_connection_message = str(exc)
        mailbox_settings.save(update_fields=['last_connection_test_at', 'last_connection_status', 'last_connection_message', 'updated_at'])
        raise
    else:
        mailbox_settings.last_connection_test_at = timezone.now()
        mailbox_settings.last_connection_status = 'success'
        mailbox_settings.last_connection_message = 'SMTP send succeeded.'
        mailbox_settings.save(update_fields=['last_connection_test_at', 'last_connection_status', 'last_connection_message', 'updated_at'])
        return True
    finally:
        client.quit()


def send_email_dispatch(dispatch):
    """Send dispatch email and persist the attachment manifest."""
    attachments = collect_email_attachments(dispatch)

    sent_via_mailbox = _send_via_user_mailbox(dispatch, attachments)
    if not sent_via_mailbox:
        email_message = DjangoEmailMessage(
            subject=dispatch.subject,
            body=dispatch.body,
            to=_parse_recipients(dispatch.sent_to),
            cc=_parse_recipients(dispatch.cc),
            bcc=_parse_recipients(dispatch.bcc),
        )

        for entry in attachments:
            path = Path(entry['path'])
            if not path.exists():
                continue
            mime_type, _ = mimetypes.guess_type(path.name)
            with open(path, 'rb') as file_obj:
                email_message.attach(path.name, file_obj.read(), mime_type or 'application/octet-stream')

        email_message.send(fail_silently=False)

    dispatch.attachments = attachments
    dispatch.sent_at = timezone.now()
    dispatch.status = EmailDispatch.DispatchStatus.SENT
    dispatch.save(update_fields=['attachments', 'sent_at', 'status', 'updated_at'])

    creator = dispatch.created_by
    mailbox_settings = getattr(creator, 'mailbox_settings', None) if creator else None
    if mailbox_settings and mailbox_settings.is_active:
        cache_dispatch_outbox_message(dispatch, mailbox_settings)

    return attachments

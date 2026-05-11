"""
Views for Products app.
"""
from datetime import date
from decimal import Decimal

from django.shortcuts import render, redirect
from django.urls import reverse_lazy
from django.contrib import messages
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, FormView, DetailView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.http import HttpResponse
from django.core.exceptions import ValidationError
from django.db import models, transaction
from django.utils import timezone

from assets.models import AssetBrand, AssetModel
from assets.models import AssetCategory
from .models import ProductPrice, ProductPriceApprovalRequest, ServiceItem
from .forms import ProductPriceForm, ServicePriceForm


class OrderManagementAccessMixin(LoginRequiredMixin, UserPassesTestMixin):
    def test_func(self):
        return self.request.user.can_manage_orders()

    def handle_no_permission(self):
        messages.error(self.request, 'You do not have access to Order Management.')
        return redirect('dashboard:dashboard')


class ProductModelCatalogMixin:
    """Provide model metadata for derived form fields and search UI."""

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        models_qs = AssetModel.objects.filter(is_active=True).exclude(
            category__item_type=AssetCategory.ItemType.SERVICE,
        ).select_related('brand').order_by('brand__name', 'name')
        context['model_catalog'] = [
            {
                'id': str(asset_model.pk),
                'brand_name': asset_model.brand.name,
                'unit': asset_model.unit or 'PCS',
                'label': str(asset_model),
                'model_number': asset_model.model_number or '',
            }
            for asset_model in models_qs
        ]
        return context


class PriceApprovalSubmissionContextMixin:
    """Expose whether the current form submits a live change or an approval request."""

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['submits_for_approval'] = not self.request.user.can_approve_order_management_prices()
        context['can_approve_order_management_prices'] = self.request.user.can_approve_order_management_prices()
        return context


class OrderManagementManagerAccessMixin(OrderManagementAccessMixin):
    """Restrict access to order-management managers and superadmins."""

    def test_func(self):
        return self.request.user.can_approve_order_management_prices()

    def handle_no_permission(self):
        messages.error(self.request, 'You do not have permission to approve or import price changes.')
        return redirect('products:price_list')


def _serialize_decimal(value):
    if value in {None, ''}:
        return ''
    return str(value)


def _serialize_date(value):
    if not value:
        return ''
    return value.isoformat()


def _serialize_id(value):
    if value in {None, ''}:
        return ''
    return str(value)


def _deserialize_decimal(value, default=None):
    if value in {None, ''}:
        return default
    return Decimal(str(value))


def _deserialize_date(value):
    if not value:
        return None
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value))


def _format_snapshot_value(value):
    if value in {None, ''}:
        return '-'
    if isinstance(value, bool):
        return 'Yes' if value else 'No'
    return value


def _validation_error_text(error):
    if hasattr(error, 'message_dict'):
        messages_list = []
        for field_name, field_errors in error.message_dict.items():
            label = field_name.replace('_', ' ').capitalize()
            messages_list.append(f"{label}: {' '.join(str(item) for item in field_errors)}")
        return ' '.join(messages_list)
    if hasattr(error, 'messages'):
        return ' '.join(str(item) for item in error.messages)
    return str(error)


def _build_price_data_snapshot(*, unit, price_without_tax, tax_rate, price_with_tax, is_current, valid_from, valid_until, notes):
    return {
        'unit': unit or '',
        'price_without_tax': _serialize_decimal(price_without_tax),
        'tax_rate': _serialize_decimal(tax_rate),
        'price_with_tax': _serialize_decimal(price_with_tax),
        'is_current': bool(is_current),
        'valid_from': _serialize_date(valid_from),
        'valid_until': _serialize_date(valid_until),
        'notes': notes or '',
    }


def _build_price_snapshot_from_instance(product_price):
    snapshot = {
        'catalog_type': product_price.catalog_type,
        'product_price': _build_price_data_snapshot(
            unit=product_price.unit,
            price_without_tax=product_price.price_without_tax,
            tax_rate=product_price.tax_rate,
            price_with_tax=product_price.price_with_tax,
            is_current=product_price.is_current,
            valid_from=product_price.valid_from,
            valid_until=product_price.valid_until,
            notes=product_price.notes,
        ),
        'display': {
            'label': product_price.display_label,
            'brand_name': product_price.display_brand_name,
            'name': product_price.display_name,
            'model_number': product_price.display_model_number,
            'unit': product_price.display_unit,
            'catalog_type': product_price.catalog_type,
        },
    }

    if product_price.is_service:
        snapshot['service_item'] = {
            'service_item_id': _serialize_id(product_price.service_item_id),
            'service_group': product_price.service_item.service_group,
            'name': product_price.service_item.name,
            'description': product_price.service_item.description,
            'unit': product_price.service_item.unit,
            'is_active': product_price.service_item.is_active,
        }
    else:
        snapshot['product_price'].update({
            'brand_id': _serialize_id(product_price.brand_id),
            'brand_name': product_price.brand.name if product_price.brand_id else '',
            'model_id': _serialize_id(product_price.model_id),
            'model_name': product_price.model.name if product_price.model_id else '',
            'model_number': product_price.model.model_number if product_price.model_id else '',
            'model_label': str(product_price.model) if product_price.model_id else '',
        })

    return snapshot


def _build_hardware_price_snapshot_from_form(form):
    model = form.cleaned_data['model']
    unit = model.unit or form.cleaned_data.get('unit') or 'PCS'
    return {
        'catalog_type': ProductPriceApprovalRequest.CatalogType.HARDWARE,
        'product_price': {
            **_build_price_data_snapshot(
                unit=unit,
                price_without_tax=form.cleaned_data.get('price_without_tax'),
                tax_rate=form.cleaned_data.get('tax_rate'),
                price_with_tax=form.cleaned_data.get('price_with_tax'),
                is_current=form.cleaned_data.get('is_current', True),
                valid_from=form.cleaned_data.get('valid_from'),
                valid_until=form.cleaned_data.get('valid_until'),
                notes=form.cleaned_data.get('notes'),
            ),
            'brand_id': _serialize_id(model.brand_id),
            'brand_name': model.brand.name,
            'model_id': _serialize_id(model.pk),
            'model_name': model.name,
            'model_number': model.model_number or '',
            'model_label': str(model),
        },
        'display': {
            'label': f'{model.brand.name} - {model.name}',
            'brand_name': model.brand.name,
            'name': model.name,
            'model_number': model.model_number or '',
            'unit': unit,
            'catalog_type': ProductPriceApprovalRequest.CatalogType.HARDWARE,
        },
    }


def _build_service_price_snapshot_from_form(form, service_item=None):
    service_group = form.cleaned_data.get('service_group', '')
    service_name = form.cleaned_data['service_name']
    unit = form.cleaned_data.get('unit') or 'JOB'
    display_label = f'{service_group} - {service_name}' if service_group else service_name
    return {
        'catalog_type': ProductPriceApprovalRequest.CatalogType.SERVICE,
        'service_item': {
            'service_item_id': _serialize_id(getattr(service_item, 'pk', None)),
            'service_group': service_group,
            'name': service_name,
            'description': form.cleaned_data.get('description', ''),
            'unit': unit,
            'is_active': True,
        },
        'product_price': _build_price_data_snapshot(
            unit=unit,
            price_without_tax=form.cleaned_data.get('price_without_tax'),
            tax_rate=form.cleaned_data.get('tax_rate'),
            price_with_tax=form.cleaned_data.get('price_with_tax'),
            is_current=form.cleaned_data.get('is_current', True),
            valid_from=form.cleaned_data.get('valid_from'),
            valid_until=form.cleaned_data.get('valid_until'),
            notes=form.cleaned_data.get('notes'),
        ),
        'display': {
            'label': display_label,
            'brand_name': service_group or 'Service',
            'name': service_name,
            'model_number': '',
            'unit': unit,
            'catalog_type': ProductPriceApprovalRequest.CatalogType.SERVICE,
        },
    }


def _build_snapshot_sections(snapshot):
    if not snapshot:
        return []

    display_data = snapshot.get('display') or {}
    product_data = snapshot.get('product_price') or {}
    service_data = snapshot.get('service_item') or {}
    sections = [
        {
            'title': 'Catalog Item',
            'rows': [
                ('Type', snapshot.get('catalog_type') or display_data.get('catalog_type')),
                ('Label', display_data.get('label')),
                ('Brand / Group', display_data.get('brand_name')),
                ('Name', display_data.get('name')),
                ('Model Number', display_data.get('model_number')),
                ('Unit', display_data.get('unit') or product_data.get('unit')),
            ],
        },
    ]

    if service_data:
        sections.append({
            'title': 'Service Item',
            'rows': [
                ('Service Group', service_data.get('service_group')),
                ('Service Name', service_data.get('name')),
                ('Description', service_data.get('description')),
                ('Unit', service_data.get('unit')),
                ('Active', service_data.get('is_active')),
            ],
        })
    else:
        sections.append({
            'title': 'Hardware Item',
            'rows': [
                ('Brand', product_data.get('brand_name')),
                ('Model', product_data.get('model_name')),
                ('Model Number', product_data.get('model_number')),
            ],
        })

    sections.append({
        'title': 'Price Details',
        'rows': [
            ('Unit', product_data.get('unit')),
            ('Price Without Tax', product_data.get('price_without_tax')),
            ('Tax Rate', product_data.get('tax_rate')),
            ('Price With Tax', product_data.get('price_with_tax')),
            ('Current Price', product_data.get('is_current')),
            ('Valid From', product_data.get('valid_from')),
            ('Valid Until', product_data.get('valid_until')),
            ('Notes', product_data.get('notes')),
        ],
    })

    for section in sections:
        section['rows'] = [
            {'label': label, 'value': _format_snapshot_value(value)}
            for label, value in section['rows']
        ]

    return sections


def _get_visible_approval_requests(user):
    queryset = ProductPriceApprovalRequest.objects.select_related(
        'requested_by',
        'reviewed_by',
        'target_price__brand',
        'target_price__model',
        'target_price__service_item',
        'target_model__brand',
        'target_service_item',
    )

    if not user.can_approve_order_management_prices():
        queryset = queryset.filter(requested_by=user)

    return queryset.annotate(
        status_priority=models.Case(
            models.When(status=ProductPriceApprovalRequest.Status.PENDING, then=models.Value(0)),
            models.When(status=ProductPriceApprovalRequest.Status.REJECTED, then=models.Value(1)),
            models.When(status=ProductPriceApprovalRequest.Status.APPROVED, then=models.Value(2)),
            default=models.Value(3),
            output_field=models.IntegerField(),
        )
    ).order_by('status_priority', '-created_at')


def _attach_latest_visible_approval_requests(product_prices, user):
    product_prices = list(product_prices)
    if not product_prices:
        return product_prices

    latest_requests_by_price_id = {}
    approval_requests = _get_visible_approval_requests(user).filter(
        target_price_id__in=[price.pk for price in product_prices],
    ).order_by('target_price_id', '-created_at', 'status_priority')

    for approval_request in approval_requests:
        latest_requests_by_price_id.setdefault(approval_request.target_price_id, approval_request)

    for product_price in product_prices:
        product_price.latest_visible_approval_request = latest_requests_by_price_id.get(product_price.pk)

    return product_prices


def _create_price_approval_request(
    *,
    request_user,
    request_type,
    catalog_type,
    target_price=None,
    target_model=None,
    target_service_item=None,
    current_snapshot=None,
    proposed_snapshot=None,
    requested_service_group='',
    requested_service_name='',
    requested_service_unit='',
    request_notes='',
):
    pending_requests = ProductPriceApprovalRequest.objects.filter(
        status=ProductPriceApprovalRequest.Status.PENDING,
    )

    if target_price is not None:
        pending_requests = pending_requests.filter(target_price=target_price)
    elif target_model is not None:
        pending_requests = pending_requests.filter(
            catalog_type=ProductPriceApprovalRequest.CatalogType.HARDWARE,
            target_model=target_model,
        )
    elif target_service_item is not None:
        pending_requests = pending_requests.filter(
            catalog_type=ProductPriceApprovalRequest.CatalogType.SERVICE,
            target_service_item=target_service_item,
        )
    elif requested_service_name:
        pending_requests = pending_requests.filter(
            catalog_type=ProductPriceApprovalRequest.CatalogType.SERVICE,
            request_type=ProductPriceApprovalRequest.RequestType.CREATE,
            requested_service_name__iexact=requested_service_name.strip(),
            requested_service_group__iexact=(requested_service_group or '').strip(),
        )

    if pending_requests.exists():
        raise ValidationError('A pending approval request already exists for this catalog item.')

    approval_request = ProductPriceApprovalRequest(
        request_type=request_type,
        catalog_type=catalog_type,
        target_price=target_price,
        target_model=target_model,
        target_service_item=target_service_item,
        requested_service_group=requested_service_group,
        requested_service_name=requested_service_name,
        requested_service_unit=requested_service_unit,
        current_snapshot=current_snapshot or {},
        proposed_snapshot=proposed_snapshot or {},
        request_notes=request_notes or '',
        requested_by=request_user,
    )
    approval_request.full_clean()
    approval_request.save()
    return approval_request


def _apply_price_data_to_instance(product_price, price_data):
    product_price.unit = price_data.get('unit') or product_price.unit or 'PCS'
    product_price.price_without_tax = _deserialize_decimal(price_data.get('price_without_tax'), Decimal('0.00'))
    product_price.tax_rate = _deserialize_decimal(price_data.get('tax_rate'), Decimal('13.00'))
    product_price.price_with_tax = _deserialize_decimal(price_data.get('price_with_tax'))
    product_price.is_current = bool(price_data.get('is_current', True))
    product_price.valid_from = _deserialize_date(price_data.get('valid_from'))
    product_price.valid_until = _deserialize_date(price_data.get('valid_until'))
    product_price.notes = price_data.get('notes') or ''


def _apply_approval_request(approval_request, reviewer, review_notes=''):
    if approval_request.status != ProductPriceApprovalRequest.Status.PENDING:
        raise ValidationError('This approval request has already been reviewed.')

    with transaction.atomic():
        if approval_request.request_type == ProductPriceApprovalRequest.RequestType.CREATE:
            proposed_snapshot = approval_request.proposed_snapshot or {}
            price_data = proposed_snapshot.get('product_price') or {}

            if approval_request.catalog_type == ProductPriceApprovalRequest.CatalogType.HARDWARE:
                model_id = price_data.get('model_id') or approval_request.target_model_id
                if not model_id:
                    raise ValidationError('The requested hardware model could not be found.')
                product_price = ProductPrice(model=AssetModel.objects.get(pk=model_id))
                _apply_price_data_to_instance(product_price, price_data)
                product_price.full_clean()
                product_price.save()
                approval_request.target_price = product_price
                approval_request.target_model = product_price.model
            else:
                service_data = proposed_snapshot.get('service_item') or {}
                service_item = approval_request.target_service_item or ServiceItem()
                service_item.service_group = service_data.get('service_group') or approval_request.requested_service_group or ''
                service_item.name = service_data.get('name') or approval_request.requested_service_name
                service_item.description = service_data.get('description') or ''
                service_item.unit = service_data.get('unit') or approval_request.requested_service_unit or price_data.get('unit') or 'JOB'
                service_item.is_active = True
                service_item.full_clean()
                service_item.save()

                product_price = ProductPrice(service_item=service_item)
                _apply_price_data_to_instance(product_price, price_data)
                product_price.service_item = service_item
                product_price.full_clean()
                product_price.save()
                approval_request.target_price = product_price
                approval_request.target_service_item = service_item

        elif approval_request.request_type == ProductPriceApprovalRequest.RequestType.UPDATE:
            if not approval_request.target_price:
                raise ValidationError('The live price for this request no longer exists.')

            proposed_snapshot = approval_request.proposed_snapshot or {}
            price_data = proposed_snapshot.get('product_price') or {}
            product_price = approval_request.target_price

            if approval_request.catalog_type == ProductPriceApprovalRequest.CatalogType.HARDWARE:
                model_id = price_data.get('model_id') or approval_request.target_model_id
                if not model_id:
                    raise ValidationError('The requested hardware model could not be found.')
                product_price.model = AssetModel.objects.get(pk=model_id)
                product_price.service_item = None
                _apply_price_data_to_instance(product_price, price_data)
                product_price.full_clean()
                product_price.save()
                approval_request.target_model = product_price.model
            else:
                service_data = proposed_snapshot.get('service_item') or {}
                service_item = product_price.service_item or approval_request.target_service_item or ServiceItem()
                service_item.service_group = service_data.get('service_group') or approval_request.requested_service_group or ''
                service_item.name = service_data.get('name') or approval_request.requested_service_name or service_item.name
                service_item.description = service_data.get('description') or ''
                service_item.unit = service_data.get('unit') or approval_request.requested_service_unit or price_data.get('unit') or 'JOB'
                service_item.is_active = True
                service_item.full_clean()
                service_item.save()

                product_price.model = None
                product_price.brand = None
                product_price.service_item = service_item
                _apply_price_data_to_instance(product_price, price_data)
                product_price.full_clean()
                product_price.save()
                approval_request.target_service_item = service_item

        else:
            if not approval_request.target_price:
                raise ValidationError('The live price for this request no longer exists.')
            if approval_request.catalog_type == ProductPriceApprovalRequest.CatalogType.HARDWARE and not approval_request.target_model_id:
                approval_request.target_model = approval_request.target_price.model
            if approval_request.catalog_type == ProductPriceApprovalRequest.CatalogType.SERVICE and not approval_request.target_service_item_id:
                approval_request.target_service_item = approval_request.target_price.service_item
            approval_request.target_price.delete()
            approval_request.target_price = None

        approval_request.status = ProductPriceApprovalRequest.Status.APPROVED
        approval_request.review_notes = review_notes or ''
        approval_request.reviewed_by = reviewer
        approval_request.reviewed_at = timezone.now()
        approval_request.save()


class ProductPriceListView(OrderManagementAccessMixin, ListView):
    """List view for product prices with filtering."""
    model = ProductPrice
    template_name = 'products/price_list.html'
    context_object_name = 'prices'
    paginate_by = 20

    def post(self, request, *args, **kwargs):
        category_id = request.POST.get('category_id')
        default_model_id = request.POST.get('default_model_id')
        category = AssetCategory.objects.filter(pk=category_id, is_active=True).first()
        if not category:
            messages.error(request, 'Selected category was not found.')
            return redirect('products:price_list')

        if default_model_id:
            model = AssetModel.objects.filter(pk=default_model_id, is_active=True, category=category).first()
            if not model:
                messages.error(request, 'Selected default model does not belong to this category.')
                return redirect('products:price_list')
            category.default_asset_model = model
            category.save(update_fields=['default_asset_model', 'updated_at'])
            messages.success(request, 'Default model updated successfully.')
        else:
            category.default_asset_model = None
            category.save(update_fields=['default_asset_model', 'updated_at'])
            messages.success(request, 'Default model cleared successfully.')
        return redirect('products:price_list')

    def get_queryset(self):
        queryset = ProductPrice.objects.all().select_related('brand', 'model', 'model__category', 'service_item')

        self.selected_type = self.request.GET.get('type') or 'all'
        if self.selected_type not in {'all', AssetCategory.ItemType.HARDWARE, AssetCategory.ItemType.SERVICE}:
            self.selected_type = 'all'

        self.selected_status = self.request.GET.get('status') or 'current'
        if self.selected_status not in {'current', 'inactive', 'all'}:
            self.selected_status = 'current'

        if self.selected_type == AssetCategory.ItemType.SERVICE:
            queryset = queryset.filter(service_item__isnull=False)
        elif self.selected_type == AssetCategory.ItemType.HARDWARE:
            queryset = queryset.filter(model__isnull=False)

        if self.selected_status == 'current':
            queryset = queryset.filter(is_current=True)
        elif self.selected_status == 'inactive':
            queryset = queryset.filter(is_current=False)

        # Filter by brand
        brand_id = self.request.GET.get('brand')
        if brand_id:
            queryset = queryset.filter(brand_id=brand_id)

        # Search by brand name or model name
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                models.Q(brand__name__icontains=search) |
                models.Q(model__name__icontains=search) |
                models.Q(model__model_number__icontains=search) |
                models.Q(model__description__icontains=search) |
                models.Q(service_item__name__icontains=search) |
                models.Q(service_item__service_group__icontains=search) |
                models.Q(service_item__description__icontains=search)
            )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        visible_approval_requests = _get_visible_approval_requests(self.request.user)
        context['prices'] = _attach_latest_visible_approval_requests(context['prices'], self.request.user)
        context['brands'] = AssetBrand.objects.filter(is_active=True).order_by('name')
        context['selected_brand'] = self.request.GET.get('brand', '')
        context['search_query'] = self.request.GET.get('search', '')
        context['selected_type'] = getattr(self, 'selected_type', 'all')
        context['selected_status'] = getattr(self, 'selected_status', 'current')
        context['can_approve_order_management_prices'] = self.request.user.can_approve_order_management_prices()
        context['pending_approval_count'] = visible_approval_requests.filter(
            status=ProductPriceApprovalRequest.Status.PENDING,
        ).count()
        categories = AssetCategory.objects.filter(
            is_active=True,
        ).exclude(
            item_type=AssetCategory.ItemType.SERVICE,
        ).prefetch_related('asset_models__brand').select_related('default_asset_model__brand').order_by('name')
        context['category_default_models'] = [
            {
                'category': category,
                'models': category.asset_models.filter(is_active=True).select_related('brand').order_by('brand__name', 'name'),
            }
            for category in categories
        ]
        return context


class ProductPriceCreateView(OrderManagementAccessMixin, PriceApprovalSubmissionContextMixin, ProductModelCatalogMixin, CreateView):
    """Create view for product prices."""
    model = ProductPrice
    form_class = ProductPriceForm
    template_name = 'products/price_form.html'
    success_url = reverse_lazy('products:price_list')

    def form_valid(self, form):
        if not self.request.user.can_approve_order_management_prices():
            try:
                _create_price_approval_request(
                    request_user=self.request.user,
                    request_type=ProductPriceApprovalRequest.RequestType.CREATE,
                    catalog_type=ProductPriceApprovalRequest.CatalogType.HARDWARE,
                    target_model=form.cleaned_data['model'],
                    proposed_snapshot=_build_hardware_price_snapshot_from_form(form),
                )
            except ValidationError as exc:
                form.add_error(None, _validation_error_text(exc))
                return self.form_invalid(form)

            messages.success(self.request, 'Product price request submitted for manager approval.')
            return redirect(self.success_url)

        messages.success(self.request, 'Product price created successfully.')
        return super().form_valid(form)


class ServicePriceCreateView(OrderManagementAccessMixin, PriceApprovalSubmissionContextMixin, FormView):
    """Create a service item and price in one step."""

    template_name = 'products/service_price_form.html'
    form_class = ServicePriceForm
    success_url = reverse_lazy('products:price_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['service_group_suggestions'] = list(
            ServiceItem.objects.exclude(service_group='').order_by('service_group').values_list('service_group', flat=True).distinct()
        )
        context['product_price'] = None
        return context

    def form_valid(self, form):
        if not self.request.user.can_approve_order_management_prices():
            try:
                _create_price_approval_request(
                    request_user=self.request.user,
                    request_type=ProductPriceApprovalRequest.RequestType.CREATE,
                    catalog_type=ProductPriceApprovalRequest.CatalogType.SERVICE,
                    proposed_snapshot=_build_service_price_snapshot_from_form(form),
                    requested_service_group=form.cleaned_data.get('service_group', ''),
                    requested_service_name=form.cleaned_data['service_name'],
                    requested_service_unit=form.cleaned_data.get('unit', 'JOB'),
                )
            except ValidationError as exc:
                form.add_error(None, exc)
                return self.form_invalid(form)

            messages.success(self.request, 'Service item request submitted for manager approval.')
            return super().form_valid(form)

        try:
            form.save()
        except ValidationError as exc:
            form.add_error(None, _validation_error_text(exc))
            return self.form_invalid(form)

        messages.success(self.request, 'Service item created successfully.')
        return super().form_valid(form)


class ServicePriceUpdateView(OrderManagementAccessMixin, PriceApprovalSubmissionContextMixin, FormView):
    """Edit a service item and its price entry."""

    template_name = 'products/service_price_form.html'
    form_class = ServicePriceForm
    success_url = reverse_lazy('products:price_list')

    def dispatch(self, request, *args, **kwargs):
        self.product_price = ProductPrice.objects.select_related('service_item').filter(
            pk=kwargs['pk'],
            service_item__isnull=False,
        ).first()
        if not self.product_price:
            messages.error(request, 'Service price was not found.')
            return redirect('products:price_list')
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['product_price'] = self.product_price
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['service_group_suggestions'] = list(
            ServiceItem.objects.exclude(service_group='').order_by('service_group').values_list('service_group', flat=True).distinct()
        )
        context['product_price'] = self.product_price
        return context

    def form_valid(self, form):
        if not self.request.user.can_approve_order_management_prices():
            try:
                _create_price_approval_request(
                    request_user=self.request.user,
                    request_type=ProductPriceApprovalRequest.RequestType.UPDATE,
                    catalog_type=ProductPriceApprovalRequest.CatalogType.SERVICE,
                    target_price=self.product_price,
                    target_service_item=self.product_price.service_item,
                    current_snapshot=_build_price_snapshot_from_instance(self.product_price),
                    proposed_snapshot=_build_service_price_snapshot_from_form(form, self.product_price.service_item),
                    requested_service_group=form.cleaned_data.get('service_group', ''),
                    requested_service_name=form.cleaned_data['service_name'],
                    requested_service_unit=form.cleaned_data.get('unit', 'JOB'),
                )
            except ValidationError as exc:
                form.add_error(None, _validation_error_text(exc))
                return self.form_invalid(form)

            messages.success(self.request, 'Service item update request submitted for manager approval.')
            return super().form_valid(form)

        try:
            form.save(product_price=self.product_price, service_item=self.product_price.service_item)
        except ValidationError as exc:
            form.add_error(None, _validation_error_text(exc))
            return self.form_invalid(form)

        messages.success(self.request, 'Service item updated successfully.')
        return super().form_valid(form)


class ProductPriceUpdateView(OrderManagementAccessMixin, PriceApprovalSubmissionContextMixin, ProductModelCatalogMixin, UpdateView):
    """Update view for product prices."""
    model = ProductPrice
    form_class = ProductPriceForm
    template_name = 'products/price_form.html'
    success_url = reverse_lazy('products:price_list')

    def form_valid(self, form):
        if not self.request.user.can_approve_order_management_prices():
            try:
                _create_price_approval_request(
                    request_user=self.request.user,
                    request_type=ProductPriceApprovalRequest.RequestType.UPDATE,
                    catalog_type=ProductPriceApprovalRequest.CatalogType.HARDWARE,
                    target_price=self.object,
                    target_model=form.cleaned_data['model'],
                    current_snapshot=_build_price_snapshot_from_instance(self.object),
                    proposed_snapshot=_build_hardware_price_snapshot_from_form(form),
                )
            except ValidationError as exc:
                form.add_error(None, _validation_error_text(exc))
                return self.form_invalid(form)

            messages.success(self.request, 'Product price update request submitted for manager approval.')
            return redirect(self.get_success_url())

        messages.success(self.request, 'Product price updated successfully.')
        return super().form_valid(form)


class ProductPriceDeleteView(OrderManagementAccessMixin, PriceApprovalSubmissionContextMixin, DeleteView):
    """Delete view for product prices."""
    model = ProductPrice
    template_name = 'products/price_confirm_delete.html'
    success_url = reverse_lazy('products:price_list')

    def form_valid(self, form):
        if not self.request.user.can_approve_order_management_prices():
            try:
                _create_price_approval_request(
                    request_user=self.request.user,
                    request_type=ProductPriceApprovalRequest.RequestType.DELETE,
                    catalog_type=(
                        ProductPriceApprovalRequest.CatalogType.SERVICE
                        if self.object.is_service else ProductPriceApprovalRequest.CatalogType.HARDWARE
                    ),
                    target_price=self.object,
                    target_model=self.object.model,
                    target_service_item=self.object.service_item,
                    current_snapshot=_build_price_snapshot_from_instance(self.object),
                    requested_service_group=(self.object.service_item.service_group if self.object.is_service else ''),
                    requested_service_name=(self.object.service_item.name if self.object.is_service else ''),
                    requested_service_unit=(self.object.service_item.unit if self.object.is_service else ''),
                )
            except ValidationError as exc:
                messages.error(self.request, _validation_error_text(exc))
                return redirect(self.success_url)

            messages.success(self.request, 'Product price delete request submitted for manager approval.')
            return redirect(self.success_url)

        messages.success(self.request, 'Product price deleted successfully.')
        return super().form_valid(form)


class ProductPriceApprovalListView(OrderManagementAccessMixin, ListView):
    """List pending and historical price approval requests."""

    model = ProductPriceApprovalRequest
    template_name = 'products/approval_request_list.html'
    context_object_name = 'approval_requests'
    paginate_by = 20

    def get_queryset(self):
        queryset = _get_visible_approval_requests(self.request.user)
        self.selected_status = self.request.GET.get('status') or ProductPriceApprovalRequest.Status.PENDING
        valid_statuses = {
            ProductPriceApprovalRequest.Status.PENDING,
            ProductPriceApprovalRequest.Status.APPROVED,
            ProductPriceApprovalRequest.Status.REJECTED,
            ProductPriceApprovalRequest.Status.CANCELLED,
            'all',
        }
        if self.selected_status not in valid_statuses:
            self.selected_status = ProductPriceApprovalRequest.Status.PENDING
        if self.selected_status != 'all':
            queryset = queryset.filter(status=self.selected_status)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['selected_status'] = getattr(self, 'selected_status', ProductPriceApprovalRequest.Status.PENDING)
        context['can_approve_order_management_prices'] = self.request.user.can_approve_order_management_prices()
        return context


class ProductPriceApprovalDetailView(OrderManagementAccessMixin, DetailView):
    """Review a single price approval request."""

    model = ProductPriceApprovalRequest
    template_name = 'products/approval_request_detail.html'
    context_object_name = 'approval_request'

    def get_queryset(self):
        return _get_visible_approval_requests(self.request.user)

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        if not request.user.can_approve_order_management_prices():
            messages.error(request, 'You do not have permission to review approval requests.')
            return redirect('products:approval_request_detail', pk=self.object.pk)

        action = request.POST.get('action')
        review_notes = (request.POST.get('review_notes') or '').strip()

        if self.object.status != ProductPriceApprovalRequest.Status.PENDING:
            messages.error(request, 'This approval request has already been reviewed.')
            return redirect('products:approval_request_detail', pk=self.object.pk)

        try:
            if action == 'approve':
                _apply_approval_request(self.object, request.user, review_notes=review_notes)
                messages.success(request, 'Approval request approved and applied successfully.')
            elif action == 'reject':
                self.object.status = ProductPriceApprovalRequest.Status.REJECTED
                self.object.review_notes = review_notes
                self.object.reviewed_by = request.user
                self.object.reviewed_at = timezone.now()
                self.object.save()
                messages.success(request, 'Approval request rejected.')
            else:
                messages.error(request, 'Select a valid review action.')
        except (AssetModel.DoesNotExist, ValidationError) as exc:
            messages.error(request, _validation_error_text(exc))

        return redirect('products:approval_request_detail', pk=self.object.pk)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['can_approve_order_management_prices'] = self.request.user.can_approve_order_management_prices()
        context['current_snapshot_sections'] = _build_snapshot_sections(self.object.current_snapshot)
        context['proposed_snapshot_sections'] = _build_snapshot_sections(self.object.proposed_snapshot)
        return context


def import_prices_view(request):
    """Import product prices from Excel file."""
    if not request.user.can_import_product_prices():
        messages.error(request, 'You do not have permission to import live product prices.')
        return redirect('products:price_list')
    from django.conf import settings
    import openpyxl
    from io import BytesIO
    import datetime

    if request.method == 'POST' and request.FILES.get('file'):
        excel_file = request.FILES['file']
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        # Get headers from first row
        headers = [cell.value for cell in sheet[1]]
        expected_headers = ['Brand_Code', 'Model_Number', 'Unit', 'Price_Without_Tax', 'Tax_Rate']

        # Validate headers
        if headers[:len(expected_headers)] != expected_headers:
            messages.error(request, f'Invalid file format. Expected headers: {", ".join(expected_headers)}')
            return redirect('products:price_import')

        success_count = 0
        error_count = 0
        errors = []

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:  # Skip empty rows
                continue

            try:
                brand_code = str(row[0]).strip()
                model_number = str(row[1]).strip() if row[1] else ''
                price_without_tax = float(row[3]) if row[3] else 0
                tax_rate = float(row[4]) if row[4] else 13.00

                # Find brand and model
                brand = AssetBrand.objects.filter(code=brand_code).first()
                if not brand:
                    errors.append(f"Row {row_num}: Brand code '{brand_code}' not found")
                    error_count += 1
                    continue

                model = brand.models.filter(model_number=model_number).first()
                if not model:
                    errors.append(f"Row {row_num}: Model number '{model_number}' not found for brand '{brand_code}'")
                    error_count += 1
                    continue

                # Create or update price
                price_with_tax = price_without_tax * (1 + tax_rate / 100)
                price, created = ProductPrice.objects.filter(is_current=True).update_or_create(
                    model=model,
                    defaults={
                        'brand': model.brand,
                        'unit': model.unit or 'PCS',
                        'price_without_tax': price_without_tax,
                        'price_with_tax': price_with_tax,
                        'tax_rate': tax_rate,
                        'is_current': True,
                    }
                )
                success_count += 1

            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
                error_count += 1

        messages.success(request, f'Import completed: {success_count} prices imported, {error_count} errors.')
        if errors:
            for error in errors[:10]:  # Show first 10 errors
                messages.warning(request, error)

        return redirect('products:price_list')

    return render(request, 'products/import_prices.html')


def download_import_template(request):
    """Download Excel template for price import."""
    if not request.user.can_import_product_prices():
        messages.error(request, 'You do not have permission to import live product prices.')
        return redirect('products:price_list')
    import openpyxl
    from io import BytesIO

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Product Prices"

    # Add headers
    headers = ['Brand_Code', 'Model_Number', 'Unit', 'Price_Without_Tax', 'Tax_Rate']
    sheet.append(headers)

    # Add example row
    sheet.append(['DELL', 'XPS-15', 'PCS', '9999.00', '13.00'])

    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        sheet.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=product_price_import_template.xlsx'
    workbook.save(response)
    return response
